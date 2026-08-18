"""
Локальная БД-зеркало (SQLite) для статистики и бэкапа.

Идея: НЕ трогаем существующую запись в Google Таблицы. Раз в N минут читаем
'Лист1' и 'Водители' целиком и полностью переливаем в SQLite (snapshot). Google
Таблица остаётся главным источником истины — база это её быстрая копия, на
которой считаем статистику без лимитов Google и держим бэкап.

Запуск вручную для проверки:
    python db.py --sync     # один снапшот из таблиц + печать статистики
    python db.py --stats    # только статистика из уже накопленной базы
"""
import os
import re
import logging
import sqlite3
from contextlib import contextmanager
from datetime import datetime, timezone, timedelta

# Тот же часовой пояс, что и в ботах (driver_bot.DUSHANBE_TZ) — фиксированный UTC+5,
# без tzdata, чтобы "сегодня" в статистике совпадало с тем, что видят боты.
DUSHANBE_TZ = timezone(timedelta(hours=5))

DB_PATH = os.getenv("SQLITE_DB_PATH", "mavsim.db")

# Английские коды статусов (как в Лист1) → человекочитаемо (RU) для вывода статистики.
STATUS_RU = {
    "NEW": "Новый",
    "READY_FOR_DRIVERS": "Свободен",
    "TAKEN": "Принял",
    "IN_TRANSIT": "В пути",
    "DELIVERED": "Доставил",
    "CANCELLED": "Отменён",
}
ACTIVE_STATUSES = ("TAKEN", "IN_TRANSIT")

_WEEKDAY_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


# ─────────────────────────── подключение / схема ───────────────────────────

_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS orders (
    order_id       TEXT PRIMARY KEY,
    status         TEXT,
    accepted_at    TEXT,   -- как в таблице: "дд.мм.гггг чч:мм"
    accepted_date  TEXT,   -- ISO "гггг-мм-дд" для GROUP BY
    delivered_at   TEXT,
    delivered_date TEXT,
    price          REAL,
    city_from      TEXT,
    city_to        TEXT,
    dtype          TEXT,
    weight         TEXT,
    courier_name   TEXT,
    courier_tg_id  TEXT,
    client_chat_id TEXT,
    source         TEXT,
    -- Поля ниже нужны не статистике, а чтению биржи/дашборда из зеркала
    -- (Фаза 1): без них из SQLite нельзя собрать карточку заказа.
    addr_from      TEXT,
    addr_to        TEXT,
    comment        TEXT,
    s_name         TEXT,
    r_phone        TEXT,
    synced_at      TEXT
);
CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status);
CREATE INDEX IF NOT EXISTS idx_orders_acc_date ON orders(accepted_date);
CREATE INDEX IF NOT EXISTS idx_orders_del_date ON orders(delivered_date);

CREATE TABLE IF NOT EXISTS drivers (
    tg_id         TEXT PRIMARY KEY,
    status        TEXT,
    registered_at TEXT,
    full_name     TEXT,
    rate          REAL,
    phone         TEXT,
    lang          TEXT,
    synced_at     TEXT
);
"""

_schema_ready = False


@contextmanager
def _connect():
    """Новое соединение на вызов — безопасно при обращении из разных потоков
    (боты дёргают snapshot через asyncio.to_thread).

    Два нюанса, на которых раньше спотыкались:
    1. Штатный контекст-менеджер sqlite3 коммитит транзакцию, но НЕ закрывает
       соединение — дескрипторы копились от вызова к вызову.
    2. Схема накатывается лениво здесь же: на свежем volume dashboard_payload()
       падал с "no such table: orders" до первого снапшота."""
    global _schema_ready
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        if not _schema_ready:
            with conn:
                conn.executescript(_SCHEMA_SQL)
                _migrate(conn)
            _schema_ready = True
        with conn:
            yield conn
    finally:
        conn.close()


def _migrate(conn) -> None:
    """Догоняет схему на уже существующей базе. `CREATE TABLE IF NOT EXISTS`
    не добавляет столбцы в таблицу, которая уже создана, — а на сервере volume
    `/data/mavsim.db` переживает пересборку контейнера, так что там лежит база
    со старой схемой. Идемпотентно: сверяем PRAGMA и добавляем недостающее."""
    have = {r["name"] for r in conn.execute("PRAGMA table_info(orders)")}
    added = [c for c in ("addr_from", "addr_to", "comment", "s_name", "r_phone") if c not in have]
    for col in added:
        conn.execute(f"ALTER TABLE orders ADD COLUMN {col} TEXT")
    if added:
        # Столбцы добавлены пустыми, а synced_at остался свежим — иначе биржа до
        # первого нового снапшота отдавала бы карточки без адресов. Помечаем
        # зеркало протухшим, чтобы is_fresh() увёл чтение в Таблицу.
        conn.execute("UPDATE orders SET synced_at = NULL")
        logging.info(f"Схема зеркала дополнена: {', '.join(added)}")


def init_db() -> None:
    """Создаёт таблицы, если их ещё нет. Идемпотентно."""
    with _connect():
        pass


# ─────────────────────────── парсеры значений ───────────────────────────

_NUM_RE = re.compile(r"-?\d+(?:[.,]\d+)?")


def _num(value) -> float:
    """Достаёт первое число из строки типа '45', '45 сом.', '1 200,50'. 0.0 если нет."""
    if value is None:
        return 0.0
    s = str(value).replace(" ", "").replace(" ", "")
    m = _NUM_RE.search(s)
    if not m:
        return 0.0
    try:
        return float(m.group(0).replace(",", "."))
    except ValueError:
        return 0.0


def _iso_date(value):
    """'дд.мм.гггг чч:мм' → 'гггг-мм-дд'. None, если не распознано/пусто."""
    s = (str(value) if value is not None else "").strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _pad(row: list, n: int) -> list:
    return (list(row) + [""] * n)[:n]


# Порядок столбцов INSERT для orders — держим в одном месте: им пользуются и
# полный снапшот, и точечный upsert из write-through.
_ORDER_COLUMNS = (
    "order_id, status, accepted_at, accepted_date, delivered_at, delivered_date, "
    "price, city_from, city_to, dtype, weight, courier_name, courier_tg_id, "
    "client_chat_id, source, addr_from, addr_to, comment, s_name, r_phone, synced_at"
)
_ORDER_PLACEHOLDERS = ",".join("?" * 21)


def _order_tuple(r: list, now_iso: str) -> tuple | None:
    """Строка «Лист1» → кортеж для INSERT в orders. None, если строка без ID
    (пустая/битая) — такие в зеркало не кладём."""
    r = _pad(r, 21)
    order_id = (r[1] or "").strip()
    if not order_id:
        return None
    return (
        order_id,
        (r[0] or "").strip(),        # status
        r[2],                         # accepted_at
        _iso_date(r[2]),              # accepted_date
        r[3],                         # delivered_at
        _iso_date(r[3]),              # delivered_date
        _num(r[4]),                   # price
        (r[5] or "").strip(),         # city_from
        (r[7] or "").strip(),         # city_to
        (r[10] or "").strip(),        # dtype
        (r[11] or "").strip(),        # weight
        (r[18] or "").strip(),        # courier_name
        (r[20] or "").strip(),        # courier_tg_id
        (r[19] or "").strip(),        # client_chat_id
        (r[17] or "").strip(),        # source
        (r[6] or "").strip(),         # addr_from
        (r[8] or "").strip(),         # addr_to
        (r[9] or "").strip(),         # comment
        (r[13] or "").strip(),        # s_name
        (r[16] or "").strip(),        # r_phone
        now_iso,
    )


# ─────────────────────────── снапшот из Google Таблиц ───────────────────────────

def snapshot_from_sheets() -> dict:
    """Читает 'Лист1' и 'Водители' целиком и ПОЛНОСТЬЮ перезаписывает таблицы в
    SQLite (mirror). Возвращает счётчики. Ничего не пишет обратно в Google.

    config импортируется лениво: чистые функции статистики не должны тянуть
    подключение к Google Таблицам."""
    import config  # lazy: подключение к Sheets происходит здесь

    init_db()
    now_iso = datetime.now(DUSHANBE_TZ).isoformat(timespec="seconds")
    counts = {"orders": 0, "drivers": 0}

    # --- Лист1 → orders ---
    if config.sheet is not None:
        rows = config.sheet.get_all_values()[1:]  # без заголовка
        order_rows = []
        for r in rows:
            tpl = _order_tuple(r, now_iso)
            if tpl:
                order_rows.append(tpl)
        with _connect() as conn:
            conn.execute("DELETE FROM orders")
            conn.executemany(
                f"INSERT INTO orders ({_ORDER_COLUMNS}) VALUES ({_ORDER_PLACEHOLDERS})",
                order_rows,
            )
        counts["orders"] = len(order_rows)

    # --- Водители → drivers ---
    if config.drivers_sheet is not None:
        rows = config.drivers_sheet.get_all_values()[1:]
        driver_rows = []
        for r in rows:
            r = _pad(r, 10)
            tg_id = (r[3] or "").strip()
            if not tg_id:
                continue
            driver_rows.append((
                tg_id,
                (r[0] or "").strip(),   # status
                r[1],                    # registered_at
                (r[2] or "").strip(),   # full_name
                _num(r[4]),              # rate
                (r[7] or "").strip(),   # phone
                (r[8] or "").strip(),   # lang
                now_iso,
            ))
        with _connect() as conn:
            conn.execute("DELETE FROM drivers")
            conn.executemany(
                """INSERT INTO drivers
                   (tg_id, status, registered_at, full_name, rate, phone, lang, synced_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                driver_rows,
            )
        counts["drivers"] = len(driver_rows)

    logging.info(f"🔄 Снапшот в SQLite: заказов {counts['orders']}, курьеров {counts['drivers']}")
    return counts


# ─────────────────── чтение для ботов (Фаза 1) ───────────────────
#
# Зачем: раньше биржа, кабинет и дашборд читали Google Таблицу напрямую, и
# любой чих Google (500/503/RemoteDisconnected) курьер видел как пустой список
# вместо заказов. Теперь читаем из зеркала, а Таблица остаётся источником
# истины для ЗАПИСИ: `_sync_take_order` и соседи по-прежнему перепроверяют
# статус в Таблице под блокировкой, поэтому устаревшая карточка в худшем
# случае даёт «заказ уже забрали», а не двойной захват.
#
# Свежесть: снапшот раз в 60с плюс write-through после каждой записи.
# Гонка «снапшот стартовал до записи и перетёр её» возможна, но самовосстановима
# — следующий снапшот вернёт правильное значение, а захват валидируется по
# Таблице в любом случае.

MIRROR_MAX_AGE_SEC = 900  # старше — считаем зеркало протухшим и идём в Таблицу


def is_fresh(max_age_sec: int = MIRROR_MAX_AGE_SEC) -> bool:
    """Есть ли в зеркале данные и достаточно ли они свежие. False → вызывающий
    код должен сходить в Google Таблицу (холодный старт, сдохший снапшот)."""
    try:
        with _connect() as conn:
            row = conn.execute("SELECT MAX(synced_at) AS ts FROM orders").fetchone()
        if not row or not row["ts"]:
            return False
        age = (datetime.now(DUSHANBE_TZ) - datetime.fromisoformat(row["ts"])).total_seconds()
        return age <= max_age_sec
    except Exception as e:
        logging.warning(f"Зеркало недоступно ({e}) — читаем из Таблицы")
        return False


def get_free_orders() -> list[dict]:
    """Свободные заказы для биржи. Форма словаря — как у driver_bot._sync_get_free_orders.
    `row_num` больше не несёт смысла (в зеркале нет номеров строк) и всё равно
    нигде не используется: действия над заказом идут по ID."""
    with _connect() as conn:
        rows = conn.execute(
            """SELECT order_id, price, city_from, city_to, addr_from, addr_to,
                      comment, dtype, s_name, r_phone
               FROM orders WHERE status = 'READY_FOR_DRIVERS'
               ORDER BY accepted_at"""
        ).fetchall()
    return [{
        "row_num":          0,
        "id":               r["order_id"],
        "price":            _price_str(r["price"]),
        "city_pickup":      r["city_from"] or "",
        "address_pickup":   r["addr_from"] or "",
        "city_delivery":    r["city_to"] or "",
        "address_delivery": r["addr_to"] or "",
        "driver_comment":   r["comment"] or "—",
        "delivery_type":    r["dtype"] or "DOOR",
        "s_name":           r["s_name"] or "—",
        "r_phone":          r["r_phone"] or "—",
    } for r in rows]


def _price_str(value) -> str:
    """В зеркале цена/ставка — число, а боты и шаблоны карточек ждут строку, как
    в Таблице. Целые печатаем без '.0'.

    Ноль отдаём ПУСТОЙ строкой, а не '0': в Таблице пустая ячейка ставки
    означает «ставка по умолчанию», и вызывающий код проверяет её на
    истинность (`float(row[4]) if row[4] else DEFAULT_DRIVER_RATE`). Строка
    '0' прошла бы эту проверку и обнулила курьеру заработок в отчёте."""
    if value is None:
        return ""
    value = float(value)
    if value == 0:
        return ""
    return str(int(value)) if value.is_integer() else str(value)


def get_orders_for_dashboard() -> tuple[list, list, list]:
    """(active, free, new) — форма как у driver_bot._sync_get_orders_for_dashboard."""
    with _connect() as conn:
        rows = conn.execute(
            """SELECT order_id, status, courier_name, courier_tg_id, city_from,
                      city_to, addr_from, addr_to, price, r_phone, s_name, accepted_at
               FROM orders
               WHERE status IN ('TAKEN','IN_TRANSIT','READY_FOR_DRIVERS','NEW')
               ORDER BY accepted_at"""
        ).fetchall()
    active, free, new = [], [], []
    for r in rows:
        status = (r["status"] or "").upper().strip()
        if status in ACTIVE_STATUSES:
            active.append({
                "row":        0,
                "id":         r["order_id"],
                "status":     status,
                "courier":    r["courier_name"] or "",
                "courier_id": r["courier_tg_id"] or "",
                "city_from":  r["city_from"] or "",
                "city_to":    r["city_to"] or "",
                "addr_from":  r["addr_from"] or "",
                "addr_to":    r["addr_to"] or "",
                "price":      _price_str(r["price"]),
                "r_phone":    r["r_phone"] or "",
                "s_name":     r["s_name"] or "",
            })
        else:
            item = {
                "row":       0,
                "id":        r["order_id"],
                "city_from": r["city_from"] or "",
                "city_to":   r["city_to"] or "",
                "price":     _price_str(r["price"]),
                "s_name":    r["s_name"] or "",
            }
            if status == "READY_FOR_DRIVERS":
                free.append(item)
            else:
                new.append({**item, "date": r["accepted_at"] or ""})
    return active, free, new


def get_drivers_for_dashboard() -> list[dict]:
    with _connect() as conn:
        rows = conn.execute(
            "SELECT tg_id, full_name FROM drivers WHERE UPPER(status) = 'ACTIVE' ORDER BY full_name"
        ).fetchall()
    return [{"fio": r["full_name"] or "", "tid": r["tg_id"], "row": 0} for r in rows]


def get_active_drivers() -> list[dict]:
    """Для авторассылки новых заказов — форма как у _sync_get_all_active_drivers."""
    with _connect() as conn:
        rows = conn.execute(
            "SELECT tg_id, full_name, lang FROM drivers WHERE UPPER(status) = 'ACTIVE'"
        ).fetchall()
    return [{
        "row_num":     0,
        "fio":         r["full_name"] or "",
        "telegram_id": r["tg_id"],
        "lang":        r["lang"] if r["lang"] in ("ru", "tj") else "ru",
    } for r in rows]


# Ширина строки листа «Водители» и позиции полей в ней — зеркало отдаёт строку
# той же формы, потому что вызывающий код индексирует её по номерам столбцов.
_DRIVER_ROW_WIDTH = 10


def get_driver_row(chat_id: str) -> list | None:
    """Строка курьера в форме листа «Водители»: [status, registered_at, fio,
    tg_id, rate, '', '', phone, lang, ''].

    Столбцы 6 (support_topic) и 9 (заявка на смену ФИО) в зеркале не хранятся —
    они нужны только редким операциям (поддержка, одобрение смены имени),
    которые продолжают ходить в Таблицу напрямую."""
    with _connect() as conn:
        r = conn.execute(
            "SELECT status, registered_at, full_name, tg_id, rate, phone, lang "
            "FROM drivers WHERE tg_id = ?", (str(chat_id),)
        ).fetchone()
    if not r:
        return None
    row = [""] * _DRIVER_ROW_WIDTH
    row[0] = r["status"] or ""
    row[1] = r["registered_at"] or ""
    row[2] = r["full_name"] or ""
    row[3] = r["tg_id"] or ""
    row[4] = _price_str(r["rate"])
    row[7] = r["phone"] or ""
    row[8] = r["lang"] or "ru"
    return row


def get_client_order_statuses(chat_id: str) -> list[dict]:
    with _connect() as conn:
        rows = conn.execute(
            "SELECT order_id, status FROM orders WHERE client_chat_id = ?", (str(chat_id),)
        ).fetchall()
    return [{"id": r["order_id"], "status": (r["status"] or "").upper().strip()} for r in rows]


# ─────────────────── write-through (Фаза 1) ───────────────────
#
# Вызывается ПОСЛЕ успешной записи в Google Таблицу, чтобы зеркало не отставало
# на целый интервал снапшота. Всё обёрнуто в try/except: зеркало — не источник
# истины, его сбой не должен ронять операцию, которая в Таблице уже прошла.

def _safe_write(sql: str, params: tuple) -> None:
    try:
        with _connect() as conn:
            conn.execute(sql, params)
    except Exception as e:
        logging.warning(f"Write-through в зеркало не удался (поправится снапшотом): {e}")


def mark_order_status(order_id: str, status: str) -> None:
    _safe_write("UPDATE orders SET status = ? WHERE order_id = ?", (status, str(order_id)))


def mark_order_taken(order_id: str, courier_name: str, courier_tg_id: str) -> None:
    _safe_write(
        "UPDATE orders SET status = 'TAKEN', courier_name = ?, courier_tg_id = ? WHERE order_id = ?",
        (courier_name, str(courier_tg_id), str(order_id)),
    )


def mark_order_released(order_id: str) -> None:
    _safe_write(
        "UPDATE orders SET status = 'READY_FOR_DRIVERS', courier_name = '', courier_tg_id = '' "
        "WHERE order_id = ?", (str(order_id),),
    )


def mark_driver_status(tg_id: str, status: str) -> None:
    _safe_write("UPDATE drivers SET status = ? WHERE tg_id = ?", (status, str(tg_id)))


def upsert_order_from_row(row: list) -> None:
    """Кладёт заказ в зеркало целиком из строки «Лист1». Нужен там, где заказа
    в зеркале может ещё не быть (менеджер отдал в работу заказ, созданный после
    последнего снапшота) — одним `UPDATE` такой не поймать."""
    try:
        tpl = _order_tuple(row, datetime.now(DUSHANBE_TZ).isoformat(timespec="seconds"))
        if not tpl:
            return
        with _connect() as conn:
            conn.execute(
                f"INSERT OR REPLACE INTO orders ({_ORDER_COLUMNS}) VALUES ({_ORDER_PLACEHOLDERS})",
                tpl,
            )
    except Exception as e:
        logging.warning(f"Write-through (upsert) в зеркало не удался: {e}")


# ─────────────────────────── статистика ───────────────────────────

def _today_iso() -> str:
    return datetime.now(DUSHANBE_TZ).date().isoformat()


def stats_overview() -> dict:
    """KPI-плитки дашборда за сегодня (по часовому поясу Душанбе)."""
    today = _today_iso()
    with _connect() as conn:
        orders_today = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE accepted_date = ?", (today,)
        ).fetchone()[0]
        delivered_today = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE status='DELIVERED' AND delivered_date = ?",
            (today,),
        ).fetchone()[0]
        in_progress = conn.execute(
            f"SELECT COUNT(*) FROM orders WHERE status IN ({','.join('?' * len(ACTIVE_STATUSES))})",
            ACTIVE_STATUSES,
        ).fetchone()[0]
        revenue_today = conn.execute(
            "SELECT COALESCE(SUM(price),0) FROM orders WHERE status='DELIVERED' AND delivered_date = ?",
            (today,),
        ).fetchone()[0]
        active_couriers = conn.execute(
            "SELECT COUNT(*) FROM drivers WHERE UPPER(status) = 'ACTIVE'"
        ).fetchone()[0]
    avg_check = round(revenue_today / delivered_today, 1) if delivered_today else 0.0
    return {
        "orders_today": orders_today,
        "delivered_today": delivered_today,
        "in_progress": in_progress,
        "revenue_today": round(revenue_today, 1),
        "avg_check": avg_check,
        "active_couriers": active_couriers,
    }


def orders_per_day(days: int = 7) -> list[dict]:
    """Число заказов по дням за последние `days` дней (включая сегодня), без пропусков дат."""
    today = datetime.now(DUSHANBE_TZ).date()
    start = today - timedelta(days=days - 1)
    with _connect() as conn:
        rows = conn.execute(
            "SELECT accepted_date, COUNT(*) c FROM orders "
            "WHERE accepted_date >= ? GROUP BY accepted_date",
            (start.isoformat(),),
        ).fetchall()
    by_date = {r["accepted_date"]: r["c"] for r in rows}
    out = []
    for i in range(days):
        d = start + timedelta(days=i)
        out.append({
            "date": d.isoformat(),
            "weekday": _WEEKDAY_RU[d.weekday()],
            "count": by_date.get(d.isoformat(), 0),
        })
    return out


def status_distribution(days: int = 30) -> list[dict]:
    """Распределение заказов по статусам за последние `days` дней (по дате принятия)."""
    start = (datetime.now(DUSHANBE_TZ).date() - timedelta(days=days - 1)).isoformat()
    with _connect() as conn:
        rows = conn.execute(
            "SELECT status, COUNT(*) c FROM orders WHERE accepted_date >= ? "
            "GROUP BY status ORDER BY c DESC",
            (start,),
        ).fetchall()
    total = sum(r["c"] for r in rows) or 1
    return [{
        "status": r["status"],
        "label": STATUS_RU.get(r["status"], r["status"] or "—"),
        "count": r["c"],
        "pct": round(r["c"] / total * 100),
    } for r in rows]


def top_couriers(limit: int = 5, days: int = 7) -> list[dict]:
    """Топ курьеров по числу доставленных заказов за последние `days` дней."""
    start = (datetime.now(DUSHANBE_TZ).date() - timedelta(days=days - 1)).isoformat()
    with _connect() as conn:
        rows = conn.execute(
            "SELECT courier_name, COUNT(*) c, COALESCE(SUM(price),0) revenue FROM orders "
            "WHERE status='DELIVERED' AND delivered_date >= ? AND courier_name <> '' "
            "GROUP BY courier_name ORDER BY c DESC LIMIT ?",
            (start, limit),
        ).fetchall()
    return [{"name": r["courier_name"], "deliveries": r["c"], "revenue": round(r["revenue"], 1)}
            for r in rows]


# ─────────────────────────── статистика по периодам ───────────────────────────

_MONTH_RU = ["", "Янв", "Фев", "Мар", "Апр", "Май", "Июн",
             "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]

# period → (человеческое имя, гранулярность, длина). Полугодие/год агрегируем по месяцам.
PERIODS = {
    "week":  {"label": "Неделя",   "gran": "day",   "days": 7},
    "month": {"label": "Месяц",    "gran": "day",   "days": 30},
    "half":  {"label": "Полгода",  "gran": "month", "months": 6},
    "year":  {"label": "Год",      "gran": "month", "months": 12},
}


def _month_floor(d):
    return d.replace(day=1)


def _add_months(d, n):
    """Прибавить n месяцев к дате (по первому числу месяца)."""
    m = d.month - 1 + n
    return d.replace(year=d.year + m // 12, month=m % 12 + 1, day=1)


def _period_range(period: str):
    """(start_date, end_date=сегодня) для периода."""
    today = datetime.now(DUSHANBE_TZ).date()
    p = PERIODS[period]
    if p["gran"] == "day":
        return today - timedelta(days=p["days"] - 1), today
    return _add_months(_month_floor(today), -(p["months"] - 1)), today


def _overview_range(conn, start: str, end: str) -> dict:
    orders = conn.execute(
        "SELECT COUNT(*) FROM orders WHERE accepted_date BETWEEN ? AND ?", (start, end)
    ).fetchone()[0]
    delivered = conn.execute(
        "SELECT COUNT(*) FROM orders WHERE status='DELIVERED' AND delivered_date BETWEEN ? AND ?",
        (start, end),
    ).fetchone()[0]
    cancelled = conn.execute(
        "SELECT COUNT(*) FROM orders WHERE status='CANCELLED' AND accepted_date BETWEEN ? AND ?",
        (start, end),
    ).fetchone()[0]
    revenue = conn.execute(
        "SELECT COALESCE(SUM(price),0) FROM orders WHERE status='DELIVERED' AND delivered_date BETWEEN ? AND ?",
        (start, end),
    ).fetchone()[0]
    avg_check = round(revenue / delivered, 1) if delivered else 0.0
    return {
        "orders": orders, "delivered": delivered, "cancelled": cancelled,
        "revenue": round(revenue, 1), "avg_check": avg_check,
    }


def _series_range(conn, start_date, end_date, gran: str) -> list[dict]:
    """Временной ряд заказов+выручки. gran='day' — по дням, 'month' — по месяцам."""
    out = []
    if gran == "day":
        rows = conn.execute(
            "SELECT accepted_date d, COUNT(*) c, "
            "COALESCE(SUM(CASE WHEN status='DELIVERED' THEN price ELSE 0 END),0) r "
            "FROM orders WHERE accepted_date BETWEEN ? AND ? GROUP BY accepted_date",
            (start_date.isoformat(), end_date.isoformat()),
        ).fetchall()
        by = {r["d"]: (r["c"], r["r"]) for r in rows}
        days = (end_date - start_date).days + 1
        for i in range(days):
            d = start_date + timedelta(days=i)
            c, r = by.get(d.isoformat(), (0, 0))
            # для недели — день недели, для месяца — число
            label = _WEEKDAY_RU[d.weekday()] if days <= 7 else str(d.day)
            out.append({"label": label, "count": c, "revenue": round(r, 1)})
    else:  # month
        rows = conn.execute(
            "SELECT substr(accepted_date,1,7) ym, COUNT(*) c, "
            "COALESCE(SUM(CASE WHEN status='DELIVERED' THEN price ELSE 0 END),0) r "
            "FROM orders WHERE accepted_date BETWEEN ? AND ? GROUP BY ym",
            (start_date.isoformat(), end_date.isoformat()),
        ).fetchall()
        by = {r["ym"]: (r["c"], r["r"]) for r in rows}
        m = start_date
        while m <= end_date:
            ym = m.strftime("%Y-%m")
            c, r = by.get(ym, (0, 0))
            out.append({"label": _MONTH_RU[m.month], "count": c, "revenue": round(r, 1)})
            m = _add_months(m, 1)
    return out


def _statuses_range(conn, start: str, end: str) -> list[dict]:
    rows = conn.execute(
        "SELECT status, COUNT(*) c FROM orders WHERE accepted_date BETWEEN ? AND ? "
        "GROUP BY status ORDER BY c DESC",
        (start, end),
    ).fetchall()
    total = sum(r["c"] for r in rows) or 1
    return [{"status": r["status"], "label": STATUS_RU.get(r["status"], r["status"] or "—"),
             "count": r["c"], "pct": round(r["c"] / total * 100)} for r in rows]


def _couriers_range(conn, start: str, end: str, limit: int = 8) -> list[dict]:
    rows = conn.execute(
        "SELECT courier_name, COUNT(*) c, COALESCE(SUM(price),0) rev FROM orders "
        "WHERE status='DELIVERED' AND delivered_date BETWEEN ? AND ? AND courier_name <> '' "
        "GROUP BY courier_name ORDER BY c DESC LIMIT ?",
        (start, end, limit),
    ).fetchall()
    return [{"name": r["courier_name"], "deliveries": r["c"], "revenue": round(r["rev"], 1)}
            for r in rows]


def dashboard_payload() -> dict:
    """Статистика по ВСЕМ периодам сразу — для кодирования в URL WebApp-дашборда.
    Страница переключает периоды на клиенте без обращения к серверу.
    Читает только из SQLite (не трогает Google Таблицы)."""
    now = datetime.now(DUSHANBE_TZ)
    with _connect() as conn:
        in_progress = conn.execute(
            f"SELECT COUNT(*) FROM orders WHERE status IN ({','.join('?' * len(ACTIVE_STATUSES))})",
            ACTIVE_STATUSES,
        ).fetchone()[0]
        active_couriers = conn.execute(
            "SELECT COUNT(*) FROM drivers WHERE UPPER(status) = 'ACTIVE'"
        ).fetchone()[0]

        periods = {}
        for key, cfg in PERIODS.items():
            start_d, end_d = _period_range(key)
            s, e = start_d.isoformat(), end_d.isoformat()
            periods[key] = {
                "label": cfg["label"],
                "overview": _overview_range(conn, s, e),
                "series": _series_range(conn, start_d, end_d, cfg["gran"]),
                "statuses": _statuses_range(conn, s, e),
                "couriers": _couriers_range(conn, s, e),
            }
    return {
        "generated_at": now.strftime("%d.%m.%Y %H:%M"),
        "now": {"in_progress": in_progress, "active_couriers": active_couriers},
        "periods": periods,
    }


# ─────────────────────────── Excel для бухгалтерии ───────────────────────────

def build_accounting_excel(period: str = "month"):
    """Реестр доставленных заказов за период для бухгалтерии.
    Возвращает (BytesIO, имя_файла, подпись). Читает только из SQLite."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if period not in PERIODS:
        period = "month"
    cfg = PERIODS[period]
    start_d, end_d = _period_range(period)
    s, e = start_d.isoformat(), end_d.isoformat()
    range_lbl = f"{start_d.strftime('%d.%m.%Y')} — {end_d.strftime('%d.%m.%Y')}"

    with _connect() as conn:
        rows = conn.execute(
            "SELECT delivered_at, delivered_date, order_id, city_from, city_to, dtype, "
            "courier_name, price FROM orders "
            "WHERE status='DELIVERED' AND delivered_date BETWEEN ? AND ? "
            "ORDER BY delivered_date, order_id",
            (s, e),
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Бухгалтерия"

    orange = PatternFill("solid", fgColor="EA580C")
    light = PatternFill("solid", fgColor="FDECE0")
    white_bold = Font(bold=True, color="FFFFFF", size=11)
    bold = Font(bold=True)
    thin = Side(style="thin", color="E5D5C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Заголовок-титул
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Реестр доставленных заказов — {cfg['label']}"
    ws["A1"].font = Font(bold=True, size=14, color="C2410C")
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Период: {range_lbl}"
    ws["A2"].font = Font(italic=True, color="6B6B6B")
    ws.row_dimensions[1].height = 22

    headers = ["Дата доставки", "ID заказа", "Город откуда", "Город куда",
               "Тип", "Курьер", "Сумма, сом."]
    hr = 4
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=col, value=h)
        c.fill = orange
        c.font = white_bold
        c.alignment = center
        c.border = border

    total = 0.0
    r = hr + 1
    for i, row in enumerate(rows):
        vals = [
            row["delivered_at"] or row["delivered_date"] or "",
            row["order_id"],
            row["city_from"], row["city_to"],
            row["dtype"], row["courier_name"],
            row["price"] or 0,
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = border
            c.alignment = center if col in (1, 2, 5, 7) else left
            if col == 7:
                c.number_format = "#,##0.00"
            if i % 2 == 1:
                c.fill = light
        total += float(row["price"] or 0)
        r += 1

    # Итоги
    ws.cell(row=r, column=6, value="ИТОГО:").font = bold
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="right")
    tc = ws.cell(row=r, column=7, value=round(total, 2))
    tc.font = bold
    tc.number_format = "#,##0.00"
    tc.fill = light
    ws.cell(row=r + 1, column=6, value="Заказов:").font = bold
    ws.cell(row=r + 1, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=r + 1, column=7, value=len(rows)).font = bold

    widths = [17, 15, 15, 15, 12, 24, 13]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"buhgalteria_{period}_{end_d.strftime('%Y%m%d')}.xlsx"
    caption = f"📑 Бухгалтерия — {cfg['label']} ({range_lbl})\nЗаказов: {len(rows)} · Сумма: {round(total, 2)} сом."
    return buf, fname, caption


# ─────────────────────────── фоновый цикл для main.py ───────────────────────────

async def run_sync_loop(interval_sec: int = 600) -> None:
    """Периодически синхронизирует базу-зеркало. Падение снапшота не должно ронять
    ботов — всё внутри try/except. Вызывается из main.py как отдельная задача."""
    import asyncio
    while True:
        try:
            await asyncio.to_thread(snapshot_from_sheets)
        except Exception as e:
            logging.error(f"Снапшот в SQLite не удался (повтор через {interval_sec}с): {e}")
        await asyncio.sleep(interval_sec)


# ─────────────────────────── CLI для проверки ───────────────────────────

def _print_stats() -> None:
    ov = stats_overview()
    print("\n📊 Сводка за сегодня:")
    print(f"  Заказов сегодня : {ov['orders_today']}")
    print(f"  Доставлено      : {ov['delivered_today']}")
    print(f"  В работе        : {ov['in_progress']}")
    print(f"  Выручка         : {ov['revenue_today']} сом.")
    print(f"  Средний чек     : {ov['avg_check']} сом.")
    print(f"  Курьеров активно: {ov['active_couriers']}")

    print("\n📈 Заказы по дням (7 дней):")
    for d in orders_per_day(7):
        bar = "█" * d["count"]
        print(f"  {d['weekday']} {d['date']}  {d['count']:>3}  {bar}")

    print("\n🥧 Статусы (30 дней):")
    for s in status_distribution(30):
        print(f"  {s['label']:<10} {s['count']:>4}  ({s['pct']}%)")

    print("\n🏆 Топ курьеров (7 дней):")
    for i, c in enumerate(top_couriers(5, 7), 1):
        print(f"  {i}. {c['name']:<20} {c['deliveries']} дост.  {c['revenue']} сом.")
    print()


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    arg = sys.argv[1] if len(sys.argv) > 1 else "--sync"
    if arg == "--stats":
        init_db()
        _print_stats()
    else:  # --sync
        snapshot_from_sheets()
        _print_stats()
