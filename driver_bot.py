import asyncio
import base64
import html
import json
import logging
import os
import threading
import traceback
from datetime import datetime, timezone, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from aiogram import types, F
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder

from config import driver_bot as bot, client_bot, manager_bot as mgr_bot, driver_dp as dp, sheet, drivers_sheet, clients_sheet, get_manager_chat_ids, sync_update_order_info_status, sync_set_delivery_time, sanitize_for_sheet, md_escape

# ─── Конфигурация ───────────────────────────────────────────────────────────
DRIVER_WEBAPP_URL   = os.getenv("DRIVER_WEBAPP_URL", "")
REPORT_PICKER_URL   = os.getenv("REPORT_PICKER_URL", "")
DEFAULT_DRIVER_RATE = float(os.getenv("DEFAULT_DRIVER_RATE", "18.0"))
LINK_TO_DRIVER_OFFER = os.getenv("DRIVER_OFFER_URL", "https://www.google.com")
SUPPORT_CHAT_ID     = os.getenv("SUPPORT_CHAT_ID", "")
NEW_ORDERS_POLL_SECONDS = int(os.getenv("NEW_ORDERS_POLL_SECONDS", "15"))
# Пауза между сообщениями в рассылке — держим отправку ниже лимита Telegram (~30 msg/s)
BROADCAST_DELAY_SECONDS = float(os.getenv("BROADCAST_DELAY_SECONDS", "0.05"))

DUSHANBE_TZ = timezone(timedelta(hours=5))

RU_MONTHS = {
    1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
    7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь",
}
RU_MONTHS_GEN = {
    1:"января",2:"февраля",3:"марта",4:"апреля",5:"мая",6:"июня",
    7:"июля",8:"августа",9:"сентября",10:"октября",11:"ноября",12:"декабря",
}

# Блокировка для атомарных операций с заказами
_order_take_lock = threading.Lock()

# Карточки заказа на бирже, разосланные курьерам — чтоб убрать "мёртвые" после захвата заказа
_job_message_refs: dict[str, list[tuple[int, int]]] = {}

# Последний показ биржи каждому курьеру — чтоб не плодить дубли при повторном нажатии
_last_jobs_msgs: dict[int, list[int]] = {}


async def _clear_previous_jobs_view(chat_id: int):
    old_ids = _last_jobs_msgs.pop(chat_id, [])
    if not old_ids:
        return
    old_id_set = set(old_ids)
    for order_id, refs in list(_job_message_refs.items()):
        remaining = [r for r in refs if not (r[0] == chat_id and r[1] in old_id_set)]
        if remaining:
            _job_message_refs[order_id] = remaining
        else:
            _job_message_refs.pop(order_id, None)
    for mid in old_ids:
        try:
            await bot.delete_message(chat_id, mid)
        except Exception:
            pass


# Последнее сообщение о статусе заявки (PENDING/"заявка отправлена") — чтоб не плодить
# при повторных /start, и стереть его как только менеджер одобрит/отклонит курьера
_status_msgs: dict[int, int] = {}


async def _replace_status_message(chat_id: int, text: str, **kwargs):
    old_id = _status_msgs.pop(chat_id, None)
    if old_id:
        try:
            await bot.delete_message(chat_id, old_id)
        except Exception:
            pass
    sent = await bot.send_message(chat_id, text, **kwargs)
    _status_msgs[chat_id] = sent.message_id
    return sent


async def _clear_status_message(chat_id: int):
    old_id = _status_msgs.pop(chat_id, None)
    if old_id:
        try:
            await bot.delete_message(chat_id, old_id)
        except Exception:
            pass


async def _try_delete(message: types.Message):
    try:
        await message.delete()
    except Exception:
        pass


async def _clear_stale_job_cards(order_id: str, keep: tuple[int, int] | None = None):
    """Стирает клавиатуру и меняет текст у всех разосланных карточек заказа, кроме той что взяли."""
    refs = _job_message_refs.pop(order_id, [])
    for chat_id, msg_id in refs:
        if keep and (chat_id, msg_id) == keep:
            continue
        try:
            driver_data = await asyncio.to_thread(_sync_get_driver, str(chat_id))
            lang = _lang_from_driver_row(driver_data)
            await bot.edit_message_text(L[lang]["job_taken_by_other"], chat_id=chat_id, message_id=msg_id)
        except Exception:
            pass


# ─── Локализация (курьерский бот — только один язык за раз) ─────────────────
JOBS_BTN     = {"ru": "🔍 Свободные заказы",              "tj": "🔍 Фармоишҳои озод"}
CABINET_BTN  = {"ru": "📊 Мой кабинет",                    "tj": "📊 Кабинети ман"}
SUPPORT_BTN  = {"ru": "📞 Поддержка",                      "tj": "📞 Дастгирӣ"}
BACK_BTN     = {"ru": "🔙 Главное меню",                   "tj": "🔙 Ба меню асосӣ"}
ACCEPT_OFFER_BTN = {"ru": "📝 Принять оферту и зарегистрироваться", "tj": "📝 Офертаро қабул кунед"}
SHARE_PHONE_BTN  = {"ru": "📱 Поделиться номером",          "tj": "📱 Рақами телефонро мубодила кунед"}
SKIP_BTN     = {"ru": "Пропустить",                        "tj": "Гузаштан"}
TAKE_JOB_BTN = {"ru": "✅ Взять заказ",                     "tj": "✅ Фармоишро гиред"}
TRANSIT_BTN  = {"ru": "🚚 В пути",                          "tj": "🚚 Дар роҳ"}
REJECT_BTN   = {"ru": "❌ Отказаться от заказа",            "tj": "❌ Аз фармоиш даст кашидан"}
DELIVERED_BTN = {"ru": "🏁 Заказ доставлен",                "tj": "🏁 Фармоиш расонида шуд"}

L = {
    "ru": {
        "lang_saved": "✅ Язык сохранён.",
        "approved": "🎉 **Поздравляем, {fio}!**\n\nВаш аккаунт курьера активирован.\nНажмите /start чтобы начать работу.",
        "rejected": (
            "❌ **Ваша заявка отклонена.**\n\n"
            "К сожалению, мы не можем принять вас на данный момент.\n"
            "По вопросам обратитесь к администрации."
        ),
        "welcome": (
            "💼 **Добро пожаловать в Mavsimi Rason!**\n\n"
            "Для работы курьером необходимо принять условия сотрудничества.\n\n"
            "📋 [Оферта для курьеров]({url})\n\n"
            "Ознакомьтесь и нажмите кнопку ниже:"
        ),
        "pending": (
            "⏳ **Заявка на рассмотрении.**\n\n"
            "Менеджер проверит ваши данные и активирует аккаунт.\n"
            "Попробуйте позже — нажмите /start чтобы проверить статус."
        ),
        "welcome_back": "👋 **С возвращением, {fio}!**\n\nВыберите действие:",
        "blocked": "⛔ Аккаунт заблокирован. Обратитесь к администратору.",
        "menu_prompt": "Выберите действие:",
        "offer_accepted_ask_fio": "✅ Отлично! Вы принимаете условия оферты.\n\nВведите ваше **ФИО** (Фамилия Имя Отчество):",
        "fio_too_short": "❌ Пожалуйста, введите полное ФИО (минимум 3 символа).",
        "fio_saved_ask_phone": "✅ Имя принято: *{fio}*\n\nТеперь введите ваш номер телефона или нажмите кнопку ниже:",
        "phone_missing": "❌ Пожалуйста, отправьте номер телефона.",
        "registered": (
            "✅ **Заявка отправлена, {fio}!**\n\n"
            "Менеджер проверит данные и активирует ваш аккаунт.\n"
            "Нажмите /start чтобы проверить статус."
        ),
        "register_error": "❌ Ошибка при регистрации. Попробуйте позже (/start).",
        "fio_empty": "❌ ФИО не может быть пустым.",
        "profile_user_not_found": "❌ Ошибка обновления. Пользователь не найден.",
        "lang_saved_fio_pending": "✅ Язык сохранён.\n✏️ Заявка на смену ФИО отправлена менеджеру — ожидайте подтверждения.",
        "bad_date": "❌ Некорректный формат даты.",
        "not_registered": "❌ Вы не зарегистрированы. Нажмите /start.",
        "generating_report": "⏳ Формирую отчёт...",
        "no_deliveries_period": "📭 За {period} доставок не найдено.",
        "report_caption": "📄 **Отчёт: {period}**\n👤 {fio}\n✅ Доставлено: {n}",
        "access_denied": "⛔ Доступ запрещён. Нажмите /start.",
        "access_denied_short": "⛔ Доступ запрещён.",
        "db_unavailable": "❌ База данных недоступна.",
        "no_free_jobs": "🕳️ На бирже сейчас нет свободных заказов.",
        "jobs_header": "📦 **Свободные заказы ({n}):**",
        "job_card": (
            "🆔 **Заказ №:** `{id}`\n"
            "• **Получатель:** {phone}\n"
            "• **Тип:** {dtype}\n"
            "• **Комментарий курьеру:** {comment}\n"
            "────────────────────\n"
            "📍 **Откуда:** {cf}, {af}\n"
            "🌆 **Куда:** {ct}, {at}\n"
            "👤 **Отправитель:** {sname}"
        ),
        "dtype_pvz": "ПВЗ 🏢", "dtype_door": "До двери 🚪",
        "new_job_header": "🆕 **Новый заказ!**\n\n",
        "job_taken_by_other": "❌ Этот заказ уже забрал другой водитель!",
        "bad_order_number": "❌ Некорректный номер заказа.",
        "order_taken": "🎉 **Вы взяли заказ {id}!**\n\nОтправляйтесь на точку забора и нажмите кнопку, когда выедете в путь.",
        "server_error": "❌ Ошибка на сервере. Попробуйте позже.",
        "reject_prompt": (
            "↩️ Отказ от заказа <b>{id}</b>\n\n"
            "Укажите причину отказа (текст или фото с подписью).\n"
            "Или нажмите «Пропустить»."
        ),
        "reject_processing": "↩️ Обрабатываю отказ...",
        "reject_failed": "❌ Не удалось отказаться — статус заказа уже изменился.",
        "reject_done": "↩️ Вы отказались от заказа <b>{id}</b>.\nЗаказ возвращён на биржу.",
        "transit_msg": "🚚 **Заказ {id}: В пути**\n\nКак доставите заказ — нажмите «Доставлен».",
        "transit_error": "❌ Ошибка при выезде. Попробуйте позже.",
        "delivered_msg": "🏁 **Заказ {id} закрыт!**\n\nОтличная работа!",
        "delivered_error": "❌ Ошибка при завершении. Попробуйте позже.",
        "support_unavailable": "⚙️ Поддержка временно недоступна.",
        "support_prompt": "📞 <b>Напишите ваш вопрос или проблему:</b>\nМы ответим в ближайшее время.",
        "support_sent": "✅ Отправлено! Менеджер ответит здесь.",
        "support_error": "❌ Ошибка. Попробуйте позже.",
        "support_session_expired": "❌ Сессия истекла. Нажмите кнопку поддержки снова.",
        "support_send_failed": "❌ Не удалось отправить. Попробуйте позже.",
        "support_reply_header": "💬 <b>Ответ от поддержки:</b>\n\n{text}",
        "support_text_only": "✍️ Пока принимаем только текст — опишите проблему сообщением.",
    },
    "tj": {
        "lang_saved": "✅ Забон нигоҳ дошта шуд.",
        "approved": "🎉 **Табрик, {fio}!**\n\nАккаунти курьери шумо фаъол шуд.\n/start-ро пахш кунед то кор оғоз кунед.",
        "rejected": (
            "❌ **Дархости шумо рад шуд.**\n\n"
            "Мутаассифона, мо дар айни ҳол шуморо қабул карда наметавонем.\n"
            "Барои саволҳо ба маъмурият муроҷиат кунед."
        ),
        "welcome": (
            "💼 **Ба Mavsimi Rason хуш омадед!**\n\n"
            "Барои кор ҳамчун курьер шартҳои ҳамкориро қабул кунед.\n\n"
            "📋 [Оферта барои курьерҳо]({url})\n\n"
            "Шартҳоро хонда, тугмаи зерро пахш кунед:"
        ),
        "pending": (
            "⏳ **Дархост дар баррасӣ.**\n\n"
            "Менеҷер маълумоти шуморо баррасӣ мекунад.\n"
            "Баъдтар /start-ро пахш кунед."
        ),
        "welcome_back": "👋 **Хуш омадед, {fio}!**\n\nАмалро интихоб кунед:",
        "blocked": "⛔ Аккаунти шумо баста шудааст.",
        "menu_prompt": "Амалро интихоб кунед:",
        "offer_accepted_ask_fio": "✅ Офертаро қабул кардед!\n\nНоми пурраи худро ворид кунед (Фамилия Ном Насаб):",
        "fio_too_short": "❌ Номи пурраи худро ворид кунед (ҳадди ақал 3 аломат).",
        "fio_saved_ask_phone": "✅ Ном қабул шуд: *{fio}*\n\nРақами телефони худро ворид кунед ё тугмаи зерро пахш кунед:",
        "phone_missing": "❌ Лутфан, рақами телефони худро фиристед.",
        "registered": (
            "✅ **Дархост фиристода шуд, {fio}!**\n\n"
            "Менеҷер маълумотро баррасӣ мекунад.\n"
            "/start-ро пахш кунед то ҳолатро тафтиш кунед."
        ),
        "register_error": "❌ Хатогӣ ҳангоми бақайдгирӣ. Баъдтар кӯшиш кунед (/start).",
        "fio_empty": "❌ Ном холӣ буда наметавонад.",
        "profile_user_not_found": "❌ Хатогӣ. Корбар дар база нест.",
        "lang_saved_fio_pending": "✅ Забон нигоҳ дошта шуд.\n✏️ Дархости тағйири ном ба менеҷер фиристода шуд — мунтазири тасдиқ шавед.",
        "bad_date": "❌ Формати сана нодуруст аст.",
        "not_registered": "❌ Шумо сабти ном нашудаед. /start-ро пахш кунед.",
        "generating_report": "⏳ Ҳисобот тайёр карда истодааст...",
        "no_deliveries_period": "📭 Дар {period} фармоише ёфт нашуд.",
        "report_caption": "📄 **Ҳисобот: {period}**\n👤 {fio}\n✅ Расонида шуд: {n}",
        "access_denied": "⛔ Дастрасӣ манъ аст. /start-ро пахш кунед.",
        "access_denied_short": "⛔ Дастрасӣ манъ аст.",
        "db_unavailable": "❌ Пойгоҳи маълумот дастнорас аст.",
        "no_free_jobs": "🕳️ Ҳоло дар бирже фармоишҳои озод нест.",
        "jobs_header": "📦 **Фармоишҳои озод ({n}):**",
        "job_card": (
            "🆔 **Фармоиш №:** `{id}`\n"
            "• **Қабулкунанда:** {phone}\n"
            "• **Намуд:** {dtype}\n"
            "• **Шарҳ барои курьер:** {comment}\n"
            "────────────────────\n"
            "📍 **Аз куҷо:** {cf}, {af}\n"
            "🌆 **Ба куҷо:** {ct}, {at}\n"
            "👤 **Фиристанда:** {sname}"
        ),
        "dtype_pvz": "ПВЗ 🏢", "dtype_door": "Ба дар 🚪",
        "new_job_header": "🆕 **Фармоиши нав!**\n\n",
        "job_taken_by_other": "❌ Ин фармоишро курьери дигар гирифт!",
        "bad_order_number": "❌ Рақами фармоиш нодуруст аст.",
        "order_taken": "🎉 **Шумо фармоиш {id} гирифтед!**\n\nБа нуқтаи забт равед ва ҳангоми баромадан ба роҳ тугмаро пахш кунед.",
        "server_error": "❌ Хатогии сервер. Баъдтар кӯшиш кунед.",
        "reject_prompt": (
            "↩️ Аз фармоиш даст кашидан <b>{id}</b>\n\n"
            "Сабаби рад карданро нависед (матн ё акс бо тавзеҳот).\n"
            "Ё «Гузаштан»-ро пахш кунед."
        ),
        "reject_processing": "↩️ Радшавӣ коркард шуда истодааст...",
        "reject_failed": "❌ Радшавӣ муяссар нашуд — статуси фармоиш тағйир ёфт.",
        "reject_done": "↩️ Шумо аз фармоиш <b>{id}</b> даст кашидед.\nФармоиш ба бирже баргашт.",
        "transit_msg": "🚚 **Фармоиш {id}: Дар роҳ**\n\nҲангоми расонидан тугмаи «Расонида шуд»-ро пахш кунед.",
        "transit_error": "❌ Хатогӣ ҳангоми баромадан ба роҳ. Баъдтар кӯшиш кунед.",
        "delivered_msg": "🏁 **Фармоиш {id} баста шуд!**\n\nКорхонаи хуб!",
        "delivered_error": "❌ Хатогӣ ҳангоми анҷом додан. Баъдтар кӯшиш кунед.",
        "support_unavailable": "⚙️ Дастгирӣ муваққатан дастнорас аст.",
        "support_prompt": "📞 <b>Саволи худро нависед:</b>\nМо ҳарчи зудтар ҷавоб хоҳем дод.",
        "support_sent": "✅ Фиристода шуд! Менеҷер ин ҷо ҷавоб хоҳад дод.",
        "support_error": "❌ Хатогӣ. Баъдтар кӯшиш кунед.",
        "support_session_expired": "❌ Мӯҳлати сессия гузашт. Тугмаи дастгириро аз нав пахш кунед.",
        "support_send_failed": "❌ Фиристода нашуд. Баъдтар кӯшиш кунед.",
        "support_reply_header": "💬 <b>Ҷавоб аз дастгирӣ:</b>\n\n{text}",
        "support_text_only": "✍️ Ҳоло танҳо матн қабул мекунем — мушкилотро бо паём нависед.",
    },
}


def _render_job_card(o: dict, lang: str, header: str = "") -> str:
    """Карточка заказа для биржи/рассылки.

    ВСЕ пользовательские поля прогоняются через md_escape: карточка уходит с
    parse_mode="Markdown", и один символ `_`/`*`/`[` в адресе или комментарии
    раньше валил отправку целиком — заказ молча не доезжал ни до кого."""
    m = md_escape
    dtype_readable = L[lang]["dtype_pvz"] if o["delivery_type"] == "PVZ" else L[lang]["dtype_door"]
    return header + L[lang]["job_card"].format(
        id=m(o["id"]),
        phone=m(o["r_phone"]),
        dtype=dtype_readable,
        comment=m(o["driver_comment"]),
        cf=m(o["city_pickup"]),
        af=m(o["address_pickup"]),
        ct=m(o["city_delivery"]),
        at=m(o["address_delivery"]),
        sname=m(o["s_name"]),
    )


class DriverRegistration(StatesGroup):
    waiting_for_lang  = State()
    waiting_for_fio   = State()
    waiting_for_phone = State()

class DriverRejectReason(StatesGroup):
    waiting_for_reason = State()

class DriverSupport(StatesGroup):
    waiting_for_message = State()
    chatting = State()


# ─── Вспомогательные функции ────────────────────────────────────────────────
async def _get_active_driver(user_id: int) -> list | None:
    data = await asyncio.to_thread(_sync_get_driver, str(user_id))
    return data if (data and data[0].upper() == "ACTIVE") else None


def _pad_row(row: list, size: int = 21) -> list:
    return row + [""] * max(0, size - len(row))


def _now_dushanbe() -> str:
    return datetime.now(DUSHANBE_TZ).strftime("%d.%m.%Y %H:%M")


def _month_label(month_str: str) -> str:
    try:
        dt = datetime.strptime(month_str, "%m.%Y")
        return f"{RU_MONTHS[dt.month]} {dt.year}"
    except ValueError:
        return month_str


def _week_label(week_start: datetime, week_end: datetime) -> str:
    if week_start.month == week_end.month:
        return f"{week_start.day}–{week_end.day} {RU_MONTHS_GEN[week_start.month]} {week_start.year}"
    return (
        f"{week_start.day} {RU_MONTHS_GEN[week_start.month]} – "
        f"{week_end.day} {RU_MONTHS_GEN[week_end.month]} {week_end.year}"
    )


def _current_week_range(now: datetime) -> tuple[datetime, datetime]:
    naive = now.replace(tzinfo=None)
    monday = naive.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=naive.weekday())
    sunday = monday + timedelta(days=6, hours=23, minutes=59, seconds=59)
    return monday, sunday


def _month_range(now: datetime) -> tuple[datetime, datetime]:
    naive = now.replace(tzinfo=None)
    first = naive.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_m = (naive.month % 12) + 1
    next_y = naive.year + (1 if naive.month == 12 else 0)
    last = datetime(next_y, next_m, 1) - timedelta(seconds=1)
    return first, last


# ─── Google Sheets: Водители ─────────────────────────────────────────────────
def _sync_get_driver(chat_id: str) -> list | None:
    if not drivers_sheet:
        return None
    try:
        cell = drivers_sheet.find(str(chat_id), in_column=4)
        return drivers_sheet.row_values(cell.row) if cell else None
    except Exception as e:
        logging.error(f"Ошибка поиска водителя {chat_id}: {e}")
        return None


def _sync_get_driver_support_topic(chat_id: str) -> str | None:
    if not drivers_sheet:
        return None
    try:
        cell = drivers_sheet.find(str(chat_id), in_column=4)
        if not cell:
            return None
        row = drivers_sheet.row_values(cell.row)
        return row[6] if len(row) > 6 and row[6] else None
    except Exception as e:
        logging.error(f"Ошибка получения driver support_topic для {chat_id}: {e}")
        return None


def _sync_save_driver_support_topic(chat_id: str, topic_id: int) -> None:
    if not drivers_sheet:
        return
    try:
        cell = drivers_sheet.find(str(chat_id), in_column=4)
        if cell:
            drivers_sheet.update_cell(cell.row, 7, str(topic_id))
    except Exception as e:
        logging.error(f"Ошибка сохранения driver support_topic для {chat_id}: {e}")


def _sync_get_driver_by_topic(topic_id: str) -> str | None:
    if not drivers_sheet:
        return None
    try:
        cell = drivers_sheet.find(str(topic_id), in_column=7)
        if not cell:
            return None
        row = drivers_sheet.row_values(cell.row)
        return row[3] if len(row) > 3 else None
    except Exception as e:
        logging.error(f"Ошибка поиска водителя по topic_id={topic_id}: {e}")
        return None


def _sync_register_driver(chat_id: str, fio: str, phone: str = "", lang: str = "ru") -> bool:
    if not drivers_sheet:
        return False
    try:
        now = _now_dushanbe()
        drivers_sheet.append_row([
            "PENDING", now, sanitize_for_sheet(fio), str(chat_id),
            str(DEFAULT_DRIVER_RATE), now, "", sanitize_for_sheet(phone), lang
        ], table_range="A1")
        return True
    except Exception as e:
        logging.error(f"Ошибка регистрации водителя {chat_id}: {e}")
        return False


def _sync_set_driver_lang(chat_id: str, lang: str) -> bool:
    """Сохраняет предпочтение языка (I) водителя по Telegram ID (D = 4). Меняется сразу, без одобрения."""
    if not drivers_sheet or lang not in ("ru", "tj"):
        return False
    try:
        cell = drivers_sheet.find(str(chat_id), in_column=4)
        if not cell:
            return False
        drivers_sheet.update_cell(cell.row, 9, lang)
        return True
    except Exception as e:
        logging.error(f"Ошибка сохранения языка водителя {chat_id}: {e}")
        return False


def _sync_request_name_change(chat_id: str, new_fio: str) -> bool:
    """Пишет заявку на смену ФИО (J) — реальное ФИО (C) не трогается до одобрения менеджером."""
    if not drivers_sheet:
        return False
    try:
        cell = drivers_sheet.find(str(chat_id), in_column=4)
        if not cell:
            return False
        drivers_sheet.update_cell(cell.row, 10, sanitize_for_sheet(new_fio))
        return True
    except Exception as e:
        logging.error(f"Ошибка заявки на смену ФИО водителя {chat_id}: {e}")
        return False


def _sync_approve_name_change(telegram_id: str) -> tuple[str, str] | None:
    """Применяет заявку на смену ФИО (J → C), очищает заявку. Returns (old_fio, new_fio) или None."""
    if not drivers_sheet:
        return None
    try:
        cell = drivers_sheet.find(str(telegram_id), in_column=4)
        if not cell:
            return None
        row = _pad_row(drivers_sheet.row_values(cell.row))
        old_fio, new_fio = row[2], row[9]
        if not new_fio:
            return None
        drivers_sheet.batch_update([
            {'range': f'C{cell.row}', 'values': [[new_fio]]},
            {'range': f'J{cell.row}', 'values': [['']]},
        ])
        return old_fio, new_fio
    except Exception as e:
        logging.error(f"Ошибка применения смены ФИО для {telegram_id}: {e}")
        return None


def _sync_reject_name_change(telegram_id: str) -> str | None:
    """Очищает заявку на смену ФИО (J), не трогая настоящее ФИО (C). Returns отклонённое ФИО или None."""
    if not drivers_sheet:
        return None
    try:
        cell = drivers_sheet.find(str(telegram_id), in_column=4)
        if not cell:
            return None
        row = _pad_row(drivers_sheet.row_values(cell.row))
        rejected_fio = row[9]
        if not rejected_fio:
            return None
        drivers_sheet.update_cell(cell.row, 10, '')
        return rejected_fio
    except Exception as e:
        logging.error(f"Ошибка отклонения смены ФИО для {telegram_id}: {e}")
        return None


def _lang_from_driver_row(row) -> str:
    """Достаёт сохранённое предпочтение языка из строки листа Водители (I = 9). По умолчанию — русский."""
    if row and len(row) > 8 and row[8] in ("ru", "tj"):
        return row[8]
    return "ru"


def _sync_get_all_active_drivers() -> list[dict]:
    if not drivers_sheet:
        return []
    try:
        all_rows = drivers_sheet.get_all_values()
        result = []
        for idx, row in enumerate(all_rows):
            if idx == 0 or len(row) < 4:
                continue
            if row[0].upper().strip() != "ACTIVE":
                continue
            result.append({"row_num": idx + 1, "fio": row[2], "telegram_id": row[3], "lang": _lang_from_driver_row(row)})
        return result
    except Exception as e:
        logging.error(f"Ошибка получения активных курьеров: {e}")
        return []


def _sync_get_driver_deliveries(chat_id: str, date_from: datetime, date_to: datetime) -> list[dict]:
    if not sheet:
        return []
    try:
        all_rows = sheet.get_all_values()
        result = []
        for idx, row in enumerate(all_rows):
            if idx == 0:
                continue
            row = _pad_row(row)
            if str(row[20]).strip() != str(chat_id):
                continue
            date_cell = row[2].strip()
            try:
                dt = datetime.strptime(date_cell[:16], "%d.%m.%Y %H:%M")
            except ValueError:
                continue
            if not (date_from <= dt <= date_to):
                continue
            tp = row[10]
            f_addr = f"{row[5]}, {row[6]}" if row[6] else row[5]
            to_addr = f"{row[7]}, {row[8]}" if tp == "DOOR" and row[8] else row[7]
            result.append({
                "i":  row[1],
                "d":  dt.strftime("%Y-%m-%d"),
                "t":  dt.strftime("%H:%M"),
                "f":  f_addr,
                "to": to_addr,
                "tp": tp,
                "s":  row[0].upper(),
            })
        return result
    except Exception as e:
        logging.error(f"Ошибка чтения доставок курьера {chat_id}: {e}")
        return []


# ─── Google Sheets: Заказы ───────────────────────────────────────────────────
def _sync_get_free_orders() -> list:
    if not sheet:
        return []
    try:
        all_rows = sheet.get_all_values()
        free_list = []
        for idx, row in enumerate(all_rows):
            if idx == 0:
                continue
            # gspread обрезает хвостовые пустые ячейки — без padding короткая строка
            # роняла построение ВСЕГО списка свободных заказов по IndexError.
            row = _pad_row(row)
            status = row[0].upper().strip()
            if status == "READY_FOR_DRIVERS":
                free_list.append({
                    "row_num":          idx + 1,
                    "id":               row[1],
                    "price":            row[4],
                    "city_pickup":      row[5],
                    "address_pickup":   row[6],
                    "city_delivery":    row[7],
                    "address_delivery": row[8],
                    "driver_comment":   row[9]  or "—",
                    "delivery_type":    row[10] or "DOOR",
                    "s_name":           row[13] or "—",
                    "r_phone":          row[16] or "—",
                })
        return free_list
    except Exception as e:
        logging.error(f"Ошибка чтения свободных заказов: {e}")
        return []


def _sync_take_order(order_id: str, courier_name: str, courier_id: str) -> bool:
    """READY_FOR_DRIVERS → TAKEN. Строка ищется по ID заказа ВНУТРИ блокировки:
    номер строки в таблице непостоянен (менеджер удаляет/вставляет строки), а
    кнопки в чатах живут долго — по номеру строки можно было забрать чужой заказ."""
    if not sheet:
        return False
    with _order_take_lock:
        try:
            found = _sync_find_order_by_id(order_id)
            if not found:
                return False
            row_num, row = found
            if row[0].upper().strip() != "READY_FOR_DRIVERS":
                return False
            sheet.batch_update([
                {"range": f"A{row_num}", "values": [["TAKEN"]]},
                {"range": f"S{row_num}", "values": [[courier_name]]},
                {"range": f"U{row_num}", "values": [[str(courier_id)]]},
            ])
            return True
        except Exception as e:
            logging.error(f"Ошибка захвата заказа {order_id}: {e}")
            return False


def _sync_release_order(order_id: str, courier_id: str) -> bool:
    """TAKEN → READY_FOR_DRIVERS. Проверяет, что именно этот курьер владеет заказом."""
    if not sheet:
        return False
    with _order_take_lock:
        try:
            found = _sync_find_order_by_id(order_id)
            if not found:
                return False
            row_num, row = found
            if row[0].upper().strip() != "TAKEN":
                return False
            if str(row[20]).strip() != str(courier_id):
                return False
            sheet.batch_update([
                {"range": f"A{row_num}", "values": [["READY_FOR_DRIVERS"]]},
                {"range": f"S{row_num}", "values": [[""]]},
                {"range": f"U{row_num}", "values": [[""]]},
            ])
            return True
        except Exception as e:
            logging.error(f"Ошибка освобождения заказа {order_id}: {e}")
            return False


def _sync_update_status(order_id: str, status: str, courier_id=None, allowed_from=None) -> bool:
    """Меняет статус заказа в Лист1 по ID. Если передан courier_id — проверяет, что
    заказ принадлежит именно этому курьеру (столбец U), а текущий статус входит
    в allowed_from. Иначе курьер мог бы двигать чужой заказ."""
    if not sheet:
        return False
    with _order_take_lock:
        try:
            found = _sync_find_order_by_id(order_id)
            if not found:
                return False
            row_num, row = found
            if courier_id is not None and str(row[20]).strip() != str(courier_id):
                return False
            if allowed_from and row[0].upper().strip() not in allowed_from:
                return False
            sheet.update_cell(row_num, 1, status)
            return True
        except Exception as e:
            logging.error(f"Ошибка обновления статуса заказа {order_id} на {status}: {e}")
            return False


def _sync_find_order_by_id(order_id: str) -> tuple[int, list] | None:
    if not sheet:
        return None
    try:
        cell = sheet.find(str(order_id), in_column=2)
        if not cell:
            return None
        return cell.row, _pad_row(sheet.row_values(cell.row))
    except Exception as e:
        logging.error(f"Ошибка поиска заказа {order_id}: {e}")
        return None


def _sync_reassign_order(order_id: str, new_courier_name: str, new_courier_id: str) -> tuple[bool, str, str]:
    """Переназначает заказ по ID. Разрешено для TAKEN/IN_TRANSIT → сбрасывает в TAKEN."""
    if not sheet:
        return False, "", ""
    with _order_take_lock:
        try:
            found = _sync_find_order_by_id(order_id)
            if not found:
                return False, "", ""
            row_num, row = found
            status = row[0].upper().strip()
            if status not in ("TAKEN", "IN_TRANSIT"):
                return False, "", ""
            old_courier_id = row[20]
            sheet.batch_update([
                {"range": f"A{row_num}", "values": [["TAKEN"]]},
                {"range": f"S{row_num}", "values": [[new_courier_name]]},
                {"range": f"U{row_num}", "values": [[str(new_courier_id)]]},
            ])
            sync_update_order_info_status(order_id, "TAKEN")
            return True, old_courier_id, order_id
        except Exception as e:
            logging.error(f"Ошибка переназначения заказа {order_id}: {e}")
            return False, "", ""


def _sync_get_orders_for_dashboard() -> tuple[list, list, list]:
    """Читает Лист1, возвращает (active_orders, free_orders, new_orders)."""
    active, free, new = [], [], []
    if not sheet:
        return active, free, new
    try:
        for idx, row in enumerate(sheet.get_all_values()):
            if idx == 0:
                continue
            row = _pad_row(row)
            status = row[0].upper().strip()
            if status in ("TAKEN", "IN_TRANSIT"):
                active.append({
                    "row":        idx + 1,
                    "id":         row[1],
                    "status":     status,
                    "courier":    row[18],
                    "courier_id": row[20],
                    "city_from":  row[5],
                    "city_to":    row[7],
                    "addr_from":  row[6],
                    "addr_to":    row[8],
                    "price":      row[4],
                    "r_phone":    row[16],
                    "s_name":     row[13],
                })
            elif status == "READY_FOR_DRIVERS":
                free.append({
                    "row":       idx + 1,
                    "id":        row[1],
                    "city_from": row[5],
                    "city_to":   row[7],
                    "price":     row[4],
                    "s_name":    row[13],
                })
            elif status == "NEW":
                new.append({
                    "row":       idx + 1,
                    "id":        row[1],
                    "city_from": row[5],
                    "city_to":   row[7],
                    "price":     row[4],
                    "s_name":    row[13],
                    "date":      row[2],
                })
    except Exception as e:
        logging.error(f"Ошибка чтения заказов для дашборда: {e}")
    return active, free, new


def _sync_get_drivers_for_dashboard() -> list:
    """Читает лист Водители, возвращает список ACTIVE курьеров."""
    result = []
    if not drivers_sheet:
        return result
    try:
        for idx, row in enumerate(drivers_sheet.get_all_values()):
            if idx == 0 or len(row) < 4:
                continue
            if row[0].upper().strip() != "ACTIVE":
                continue
            result.append({"fio": row[2], "tid": row[3], "row": idx + 1})
    except Exception as e:
        logging.error(f"Ошибка чтения курьеров для дашборда: {e}")
    return result


async def _async_get_admin_dashboard_data() -> dict:
    """Читает оба листа параллельно — вдвое быстрее последовательного чтения."""
    (active, free, new), drivers = await asyncio.gather(
        asyncio.to_thread(_sync_get_orders_for_dashboard),
        asyncio.to_thread(_sync_get_drivers_for_dashboard),
    )
    busy_ids = {o["courier_id"] for o in active}
    couriers = [{**d, "busy": d["tid"] in busy_ids} for d in drivers]
    return {"orders": active, "free": free, "new": new, "couriers": couriers}


# ─── Excel-отчёт ─────────────────────────────────────────────────────────────
def generate_excel_report(driver_name: str, rate: float, deliveries: list[dict], period_label: str,
                           show_earnings: bool = True) -> BytesIO:
    """show_earnings=False скрывает ставку/заработок — используется для отчёта, который скачивает сам курьер."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    n_cols = 8 if show_earnings else 7

    BRAND   = "FFEA580C"
    L_BRAND = "FFFDECE0"
    WHITE  = "FFFFFFFF"
    GRAY   = "FFF4F4F5"

    thin = Side(border_style="thin", color="FFD0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(ws, row, col, value, bold=False, bg=None, color="FF1C1C1E",
                   align="left", size=11, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=color, size=size, name="Calibri")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border = border
        return c

    last_col_letter = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws.cell(row=1, column=1, value="MAVSIMI RASON — Отчёт курьера")
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.fill = PatternFill("solid", fgColor=BRAND)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"A2:{last_col_letter}2")
    info_line = f"Курьер: {driver_name}   |   Период: {period_label}"
    if show_earnings:
        info_line += f"   |   Ставка: {rate:.2f} TJS / доставка"
    c2 = ws.cell(row=2, column=1, value=info_line)
    c2.font = Font(bold=False, color="FF555555", size=11, name="Calibri")
    c2.fill = PatternFill("solid", fgColor=GRAY)
    c2.alignment = Alignment(horizontal="center", vertical="center")

    delivered   = [d for d in deliveries if d["s"] == "DELIVERED"]
    total_count = len(delivered)
    total_earn  = total_count * rate

    if show_earnings:
        ws.merge_cells("A3:D3")
        ws.merge_cells("E3:H3")
    else:
        ws.merge_cells(f"A3:{last_col_letter}3")
    c3a = ws.cell(row=3, column=1, value=f"Итого доставок: {total_count}")
    c3a.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
    c3a.fill = PatternFill("solid", fgColor=BRAND)
    c3a.alignment = Alignment(horizontal="center", vertical="center")
    if show_earnings:
        c3b = ws.cell(row=3, column=5, value=f"Итого к выплате: {total_earn:.2f} TJS")
        c3b.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
        c3b.fill = PatternFill("solid", fgColor="FF1F9D4D")
        c3b.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["№", "Дата", "Время", "Откуда", "Куда", "Тип", "Статус"]
    if show_earnings:
        headers.append("Заработок (TJS)")
    for col, h in enumerate(headers, 1):
        cell_style(ws, 4, col, h, bold=True, bg=BRAND, color=WHITE, align="center", size=11)

    col_widths = [5, 13, 9, 18, 18, 12, 14] + ([16] if show_earnings else [])
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 20

    for i, d in enumerate(deliveries, 1):
        r = i + 4
        bg = L_BRAND if i % 2 == 0 else WHITE
        is_done = d["s"] == "DELIVERED"
        earn = rate if is_done else 0.0
        status_label = {
            "DELIVERED": "✓ Доставлен", "IN_TRANSIT": "В пути", "TAKEN": "Принял",
        }.get(d["s"], d["s"])
        cell_style(ws, r, 1, i,            bg=bg, align="center")
        cell_style(ws, r, 2, d["d"],       bg=bg, align="center")
        cell_style(ws, r, 3, d["t"],       bg=bg, align="center")
        cell_style(ws, r, 4, d["f"],       bg=bg)
        cell_style(ws, r, 5, d["to"],      bg=bg)
        cell_style(ws, r, 6, "ПВЗ" if d["tp"] == "PVZ" else "До двери", bg=bg, align="center")
        cell_style(ws, r, 7, status_label, bg=bg, align="center",
                   color="FF1F9D4D" if is_done else "FF555555", bold=is_done)
        if show_earnings:
            earn_cell = cell_style(ws, r, 8, earn, bg=bg, align="center",
                                   bold=is_done, color="FFEA580C" if is_done else "FF999999")
            earn_cell.number_format = '0.00 "TJS"'
        ws.row_dimensions[r].height = 18

    last = len(deliveries) + 5
    if show_earnings:
        ws.merge_cells(f"A{last}:G{last}")
        cell_style(ws, last, 1, "ИТОГО К ВЫПЛАТЕ:", bold=True, bg=GRAY, align="right", size=12)
        earn_total = ws.cell(row=last, column=8, value=total_earn)
        earn_total.font = Font(bold=True, color="FF1F9D4D", size=13, name="Calibri")
        earn_total.fill = PatternFill("solid", fgColor=GRAY)
        earn_total.alignment = Alignment(horizontal="center", vertical="center")
        earn_total.number_format = '0.00 "TJS"'
        earn_total.border = border
        ws.row_dimensions[last].height = 22
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Клавиатура главного меню водителя ──────────────────────────────────────
async def build_driver_main_menu(driver_id: int):
    driver_data = await asyncio.to_thread(_sync_get_driver, str(driver_id))
    lang = _lang_from_driver_row(driver_data)
    b = ReplyKeyboardBuilder()
    b.button(text=JOBS_BTN[lang])
    if DRIVER_WEBAPP_URL:
        try:
            if driver_data and driver_data[0].upper() == "ACTIVE":
                fio  = driver_data[2] if len(driver_data) > 2 else "Курьер"
                rate = float(driver_data[4]) if len(driver_data) > 4 and driver_data[4] else DEFAULT_DRIVER_RATE
                now = datetime.now(DUSHANBE_TZ)
                date_from, date_to = _month_range(now)
                week_start_dt, week_end_dt = _current_week_range(now)
                deliveries = await asyncio.to_thread(
                    _sync_get_driver_deliveries, str(driver_id), date_from, date_to
                )
                payload = {
                    "name": fio, "rate": rate,
                    "month": now.strftime("%m.%Y"),
                    "month_label": _week_label(week_start_dt, week_end_dt),
                    "deliveries": deliveries,
                    "lang": lang,
                }
                b64 = base64.urlsafe_b64encode(
                    json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode()
                ).decode().rstrip("=")
                b.button(text=CABINET_BTN[lang], web_app=types.WebAppInfo(url=f"{DRIVER_WEBAPP_URL}?d={b64}"))
            else:
                b.button(text=CABINET_BTN[lang])
        except Exception:
            b.button(text=CABINET_BTN[lang])
    else:
        b.button(text=CABINET_BTN[lang])
    b.button(text=SUPPORT_BTN[lang])
    b.adjust(1)
    return b.as_markup(resize_keyboard=True)


def _sync_client_lang(chat_id: str) -> str:
    """Язык клиента (лист Клиенты, H=8) по Chat ID (F=6). Дефолт 'ru'."""
    if not clients_sheet or not str(chat_id).strip():
        return "ru"
    try:
        cell = clients_sheet.find(str(chat_id), in_column=6)
        if not cell:
            return "ru"
        row = clients_sheet.row_values(cell.row)
        return row[7] if len(row) > 7 and row[7] in ("ru", "tj") else "ru"
    except Exception as e:
        logging.error(f"Ошибка чтения языка клиента {chat_id}: {e}")
        return "ru"


async def send_client_push(chat_id: str, ru: str, tj: str | None = None):
    """Шлёт пуш клиенту на ЕГО языке (ru или tj), а не оба сразу.
    tj опционален: если не передан (напр. пуши из manager_bot без перевода),
    клиенту уходит русский текст — лучше, чем TypeError и молчание."""
    if not (client_bot and chat_id and str(chat_id).isdigit()):
        return
    lang = await asyncio.to_thread(_sync_client_lang, chat_id)
    try:
        await client_bot.send_message(
            chat_id=int(chat_id),
            text=(tj if lang == "tj" and tj else ru),
            parse_mode="Markdown",
        )
    except Exception as e:
        logging.error(f"Не удалось отправить пуш клиенту {chat_id}: {e}")


# ─── Глобальная навигация ────────────────────────────────────────────────────
@dp.message(F.text.in_({BACK_BTN["ru"], BACK_BTN["tj"]}))
async def driver_go_main_menu(message: types.Message, state: FSMContext):
    await state.clear()
    await _try_delete(message)
    driver_data = await asyncio.to_thread(_sync_get_driver, str(message.from_user.id))
    lang = _lang_from_driver_row(driver_data)
    await message.answer(L[lang]["menu_prompt"], reply_markup=await build_driver_main_menu(message.from_user.id))


# ─── Регистрация ─────────────────────────────────────────────────────────────
@dp.message(CommandStart())
async def cmd_start_driver(message: types.Message, state: FSMContext):
    await state.clear()
    await _try_delete(message)
    driver_data = await asyncio.to_thread(_sync_get_driver, str(message.from_user.id))

    if not driver_data:
        b = InlineKeyboardBuilder()
        b.button(text="🇷🇺 Русский", callback_data="reglang:ru")
        b.button(text="🇹🇯 Тоҷикӣ", callback_data="reglang:tj")
        b.adjust(2)
        await _replace_status_message(
            message.chat.id, "🌐 Выберите язык / Забонро интихоб кунед:", reply_markup=b.as_markup()
        )
        await state.set_state(DriverRegistration.waiting_for_lang)
    elif driver_data[0].upper() == "PENDING":
        lang = _lang_from_driver_row(driver_data)
        await _replace_status_message(message.chat.id, L[lang]["pending"], parse_mode="Markdown")
    elif driver_data[0].upper() == "ACTIVE":
        fio = driver_data[2] if len(driver_data) > 2 else "Курьер"
        lang = _lang_from_driver_row(driver_data)
        await _clear_status_message(message.chat.id)
        await message.answer(
            L[lang]["welcome_back"].format(fio=fio),
            reply_markup=await build_driver_main_menu(message.from_user.id),
            parse_mode="Markdown"
        )
    else:
        lang = _lang_from_driver_row(driver_data)
        await message.answer(L[lang]["blocked"])


@dp.callback_query(F.data.startswith("reglang:"), DriverRegistration.waiting_for_lang)
async def set_registration_lang(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    lang = callback.data.split(":")[1]
    if lang not in ("ru", "tj"):
        lang = "ru"
    await state.update_data(reg_lang=lang)
    b = ReplyKeyboardBuilder()
    b.button(text=ACCEPT_OFFER_BTN[lang])
    await _replace_status_message(
        callback.message.chat.id,
        L[lang]["welcome"].format(url=LINK_TO_DRIVER_OFFER),
        reply_markup=b.as_markup(resize_keyboard=True),
        parse_mode="Markdown",
        disable_web_page_preview=True
    )


@dp.message(F.text.in_({ACCEPT_OFFER_BTN["ru"], ACCEPT_OFFER_BTN["tj"]}))
async def accept_offer(message: types.Message, state: FSMContext):
    await _try_delete(message)
    data = await state.get_data()
    lang = data.get("reg_lang", "ru")
    await _replace_status_message(
        message.chat.id,
        L[lang]["offer_accepted_ask_fio"],
        reply_markup=types.ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )
    await state.update_data(reg_lang=lang)
    await state.set_state(DriverRegistration.waiting_for_fio)


@dp.message(DriverRegistration.waiting_for_fio)
async def save_driver_fio(message: types.Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("reg_lang", "ru")
    fio = message.text.strip() if message.text else ""
    if not fio or len(fio) < 3:
        await message.answer(L[lang]["fio_too_short"])
        return
    await _try_delete(message)
    await state.update_data(fio=fio)
    await state.set_state(DriverRegistration.waiting_for_phone)
    b = ReplyKeyboardBuilder()
    b.button(text=SHARE_PHONE_BTN[lang], request_contact=True)
    await _replace_status_message(
        message.chat.id,
        L[lang]["fio_saved_ask_phone"].format(fio=fio),
        reply_markup=b.as_markup(resize_keyboard=True, one_time_keyboard=True),
        parse_mode="Markdown"
    )


@dp.message(DriverRegistration.waiting_for_phone)
async def save_driver_phone(message: types.Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("reg_lang", "ru")
    if message.contact:
        phone = message.contact.phone_number
        if not phone.startswith("+"):
            phone = "+" + phone
    elif message.text:
        phone = message.text.strip()
    else:
        await message.answer(L[lang]["phone_missing"])
        return

    await _try_delete(message)
    fio = data.get("fio", "")
    await state.clear()

    success = await asyncio.to_thread(_sync_register_driver, str(message.from_user.id), fio, phone, lang)
    if success:
        await _replace_status_message(
            message.chat.id,
            L[lang]["registered"].format(fio=fio),
            reply_markup=types.ReplyKeyboardRemove(),
            parse_mode="Markdown"
        )
        if mgr_bot:
            manager_ids = await asyncio.to_thread(get_manager_chat_ids)
            if manager_ids:
                b = InlineKeyboardBuilder()
                b.button(text="✅ Одобрить", callback_data=f"approve_driver:{message.from_user.id}")
                b.button(text="❌ Отклонить", callback_data=f"reject_driver:{message.from_user.id}")
                b.adjust(2)
                mgr_text = (
                    f"👤 <b>Новый курьер</b>\n"
                    f"ФИО: <b>{html.escape(fio)}</b>\n"
                    f"📱 Телефон: <code>{html.escape(phone)}</code>\n"
                    f"ID: <code>{message.from_user.id}</code>\n\n"
                    f"Одобрить заявку?"
                )
                for mgr_id in manager_ids:
                    try:
                        await mgr_bot.send_message(
                            chat_id=int(mgr_id),
                            text=mgr_text,
                            reply_markup=b.as_markup(),
                            parse_mode="HTML"
                        )
                    except Exception as e:
                        logging.error(f"Не удалось уведомить менеджера {mgr_id} о новом курьере: {e}")
    else:
        await message.answer(L[lang]["register_error"])


@dp.message(F.web_app_data)
async def handle_webapp(message: types.Message):
    try:
        data = json.loads(message.web_app_data.data)
    except Exception:
        return

    action = data.get("action")

    # ── Обновление профиля (язык — сразу, ФИО — заявка менеджеру) ────────────
    if action == "update_profile":
        updated_fio = data.get("fio", "").strip()
        new_lang = data.get("lang") if data.get("lang") in ("ru", "tj") else None

        driver_data = await asyncio.to_thread(_sync_get_driver, str(message.from_user.id))
        lang = new_lang or _lang_from_driver_row(driver_data)
        if not updated_fio:
            await message.answer(L[lang]["fio_empty"])
            return
        if not driver_data:
            await message.answer(L[lang]["profile_user_not_found"])
            return
        current_fio = driver_data[2] if len(driver_data) > 2 else ""

        if new_lang:
            await asyncio.to_thread(_sync_set_driver_lang, str(message.from_user.id), new_lang)

        fio_changed = updated_fio != current_fio
        if fio_changed:
            await asyncio.to_thread(_sync_request_name_change, str(message.from_user.id), updated_fio)
            if mgr_bot:
                manager_ids = await asyncio.to_thread(get_manager_chat_ids)
                if manager_ids:
                    nb = InlineKeyboardBuilder()
                    nb.button(text="✅ Одобрить", callback_data=f"napprove:{message.from_user.id}")
                    nb.button(text="❌ Отклонить", callback_data=f"nreject:{message.from_user.id}")
                    nb.adjust(2)
                    nb_text = (
                        f"✏️ <b>Заявка на смену ФИО курьера</b>\n"
                        f"Было: {html.escape(current_fio)}\n"
                        f"Стало: {html.escape(updated_fio)}\n"
                        f"ID: <code>{message.from_user.id}</code>"
                    )
                    for mgr_id in manager_ids:
                        try:
                            await mgr_bot.send_message(
                                chat_id=int(mgr_id),
                                text=nb_text,
                                reply_markup=nb.as_markup(),
                                parse_mode="HTML"
                            )
                        except Exception as e:
                            logging.error(f"Не удалось уведомить менеджера {mgr_id} о смене ФИО: {e}")

        driver_data = await asyncio.to_thread(_sync_get_driver, str(message.from_user.id))
        lang = _lang_from_driver_row(driver_data)
        msg = L[lang]["lang_saved_fio_pending"] if fio_changed else L[lang]["lang_saved"]
        await message.answer(msg, reply_markup=await build_driver_main_menu(message.from_user.id), parse_mode="Markdown")
        return

    # ── Еженедельный отчёт ───────────────────────────────────────────────────
    if action == "generate_report":
        if not await _get_active_driver(message.from_user.id):
            return
        driver_data = await asyncio.to_thread(_sync_get_driver, str(message.from_user.id))
        if not driver_data:
            await message.answer(L["ru"]["not_registered"])
            return
        lang = _lang_from_driver_row(driver_data)

        week_start_str = data.get("week_start", "")
        try:
            week_start = datetime.strptime(week_start_str, "%Y-%m-%d")
        except ValueError:
            await message.answer(L[lang]["bad_date"])
            return
        week_end = week_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
        period_label = _week_label(week_start, week_end)

        fio = driver_data[2] if len(driver_data) > 2 else "Курьер"
        try:
            rate = float(driver_data[4]) if len(driver_data) > 4 and driver_data[4] else DEFAULT_DRIVER_RATE
        except (ValueError, TypeError):
            rate = DEFAULT_DRIVER_RATE

        wait_msg = await message.answer(L[lang]["generating_report"])
        deliveries = await asyncio.to_thread(
            _sync_get_driver_deliveries, str(message.from_user.id), week_start, week_end
        )

        if not deliveries:
            await wait_msg.delete()
            await message.answer(L[lang]["no_deliveries_period"].format(period=period_label))
            return

        excel_buf = await asyncio.to_thread(
            generate_excel_report, fio, rate, deliveries, period_label, show_earnings=False
        )
        filename = f"report_{week_start_str}_{message.from_user.id}.xlsx"
        delivered_count = sum(1 for d in deliveries if d["s"] == "DELIVERED")
        await wait_msg.delete()
        await message.answer_document(
            types.BufferedInputFile(excel_buf.read(), filename=filename),
            caption=L[lang]["report_caption"].format(period=period_label, fio=fio, n=delivered_count),
            parse_mode="Markdown"
        )
        return

    logging.warning(f"Неизвестный action из WebApp: {action}")


# ─── Биржа заказов ───────────────────────────────────────────────────────────
@dp.message(F.text.in_({JOBS_BTN["ru"], JOBS_BTN["tj"]}))
async def show_jobs(message: types.Message):
    driver_data = await _get_active_driver(message.from_user.id)
    if not driver_data:
        await message.answer(L["ru"]["access_denied"])
        return
    lang = _lang_from_driver_row(driver_data)
    chat_id = message.chat.id
    await _try_delete(message)
    await _clear_previous_jobs_view(chat_id)

    try:
        free_jobs = await asyncio.to_thread(_sync_get_free_orders)
        if not free_jobs:
            sent = await message.answer(L[lang]["no_free_jobs"])
            _last_jobs_msgs[chat_id] = [sent.message_id]
            return

        msg_ids = []
        header = await message.answer(L[lang]["jobs_header"].format(n=len(free_jobs)), parse_mode="Markdown")
        msg_ids.append(header.message_id)
        for o in free_jobs:
            b = InlineKeyboardBuilder()
            b.button(text=TAKE_JOB_BTN[lang], callback_data=f"take:{o['id']}")
            try:
                sent = await message.answer(
                    _render_job_card(o, lang), reply_markup=b.as_markup(), parse_mode="Markdown"
                )
            except Exception as e:
                # одна битая карточка не должна лишать курьера всей биржи
                logging.error(f"Не удалось показать карточку заказа {o['id']}: {e}")
                continue
            _job_message_refs.setdefault(o['id'], []).append((chat_id, sent.message_id))
            msg_ids.append(sent.message_id)
        _last_jobs_msgs[chat_id] = msg_ids
    except Exception:
        logging.error(f"Сбой показа биржи заказов: {traceback.format_exc()}")
        await message.answer(L[lang]["server_error"])


# ─── Управление заказом ──────────────────────────────────────────────────────
@dp.callback_query(F.data.startswith("take:"))
async def accept_order(callback: types.CallbackQuery):
    await callback.answer()
    driver_data = await _get_active_driver(callback.from_user.id)
    if not driver_data:
        await callback.message.answer(L["ru"]["access_denied"])
        return
    lang = _lang_from_driver_row(driver_data)
    try:
        order_id = callback.data.split(":", 1)[1]
        if not order_id or not sheet:
            await callback.message.answer(L[lang]["bad_order_number"])
            return

        c_name = driver_data[2] if len(driver_data) > 2 and driver_data[2] else callback.from_user.full_name
        c_id   = callback.from_user.id

        success = await asyncio.to_thread(_sync_take_order, order_id, c_name, c_id)
        if not success:
            await callback.message.edit_text(L[lang]["job_taken_by_other"], reply_markup=None)
            return
        await asyncio.to_thread(sync_update_order_info_status, order_id, "TAKEN")

        found = await asyncio.to_thread(_sync_find_order_by_id, order_id)
        client_chat_id = found[1][19] if found else ""

        b = InlineKeyboardBuilder()
        b.button(text=TRANSIT_BTN[lang], callback_data=f"transit:{order_id}")
        b.button(text=REJECT_BTN[lang], callback_data=f"reject:{order_id}")
        b.adjust(1)
        await callback.message.edit_text(
            L[lang]["order_taken"].format(id=order_id),
            reply_markup=b.as_markup(), parse_mode="Markdown"
        )
        await _clear_stale_job_cards(order_id, keep=(callback.from_user.id, callback.message.message_id))
        tracked = _last_jobs_msgs.get(callback.from_user.id)
        if tracked and callback.message.message_id in tracked:
            tracked.remove(callback.message.message_id)
        if client_chat_id:
            safe_name = md_escape(c_name)
            await send_client_push(client_chat_id,
                ru=f"🚚 **Ваш заказ {order_id} принят курьером!**\n👤 **Курьер:** {safe_name}",
                tj=f"🚚 **Фармоиши шумо {order_id} қабул шуд!**\n👤 **Курьер:** {safe_name}")
    except Exception:
        logging.error(f"Сбой take: {traceback.format_exc()}")
        await callback.message.answer(L[lang]["server_error"])


@dp.callback_query(F.data.startswith("reject:"))
async def reject_order(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    driver_data = await _get_active_driver(callback.from_user.id)
    if not driver_data:
        await callback.message.answer(L["ru"]["access_denied_short"])
        return
    lang = _lang_from_driver_row(driver_data)
    try:
        order_id = callback.data.split(":", 1)[1]
        if not sheet:
            await callback.message.answer(L[lang]["db_unavailable"])
            return
        found = await asyncio.to_thread(_sync_find_order_by_id, order_id)
        if not found:
            await callback.message.answer(L[lang]["bad_order_number"])
            return
        client_chat_id = found[1][19]

        await state.set_state(DriverRejectReason.waiting_for_reason)
        await state.update_data(
            order_id=order_id,
            client_chat_id=client_chat_id,
            courier_name=driver_data[2] if len(driver_data) > 2 and driver_data[2] else callback.from_user.full_name,
            courier_id=str(callback.from_user.id),
            lang=lang,
        )

        b = InlineKeyboardBuilder()
        b.button(text=SKIP_BTN[lang], callback_data="reject_skip")
        await callback.message.edit_text(
            L[lang]["reject_prompt"].format(id=order_id),
            reply_markup=b.as_markup(),
            parse_mode="HTML"
        )
    except Exception:
        logging.error(f"Сбой reject: {traceback.format_exc()}")
        await callback.message.answer(L[lang]["server_error"])


async def _do_reject(chat_id: int, state: FSMContext, reason: str | None, photo_file_id: str | None):
    """Финализирует отказ: освобождает заказ, уведомляет клиента и менеджера."""
    data = await state.get_data()
    await state.clear()

    order_id       = data["order_id"]
    client_chat_id = data["client_chat_id"]
    c_name         = data["courier_name"]
    c_id           = data["courier_id"]
    lang           = data.get("lang", "ru")

    success = await asyncio.to_thread(_sync_release_order, order_id, c_id)
    if not success:
        from config import driver_bot as _bot
        await _bot.send_message(chat_id, L[lang]["reject_failed"])
        return
    await asyncio.to_thread(sync_update_order_info_status, order_id, "READY_FOR_DRIVERS")

    from config import driver_bot as _bot
    await _bot.send_message(
        chat_id,
        L[lang]["reject_done"].format(id=order_id),
        parse_mode="HTML"
    )

    async def _notify_client():
        if not client_chat_id:
            return
        await send_client_push(
            client_chat_id,
            ru=f"ℹ️ По вашему заказу *{order_id}* происходят изменения — ищем нового курьера.",
            tj=f"ℹ️ Дар фармоиши шумо *{order_id}* тағйирот рӯй дод — курьери нав меҷӯем."
        )

    async def _notify_manager():
        if not mgr_bot:
            return
        manager_ids = await asyncio.to_thread(get_manager_chat_ids)
        if not manager_ids:
            return
        mgr_text = (
            f"⚠️ Курьер <b>{html.escape(c_name)}</b> отказался от заказа <b>{order_id}</b>.\n"
            f"📝 Причина: {html.escape(reason) if reason else '—'}"
        )
        photo_bytes = None
        if photo_file_id:
            # file_id привязан к боту, который его получил (driver_bot) —
            # manager_bot чужой file_id использовать не может, скачиваем и грузим заново
            photo_bytes = (await bot.download(photo_file_id)).read()
        for mgr_id in manager_ids:
            try:
                if photo_bytes:
                    await mgr_bot.send_photo(
                        chat_id=int(mgr_id),
                        photo=types.BufferedInputFile(photo_bytes, filename="reject.jpg"),
                        caption=mgr_text,
                        parse_mode="HTML"
                    )
                else:
                    await mgr_bot.send_message(
                        chat_id=int(mgr_id),
                        text=mgr_text,
                        parse_mode="HTML"
                    )
            except Exception as e:
                logging.error(f"Не удалось уведомить менеджера {mgr_id} об отказе от заказа {order_id}: {e}")

    results = await asyncio.gather(_notify_client(), _notify_manager(), return_exceptions=True)
    for label, result in zip(("клиента", "менеджера"), results):
        if isinstance(result, Exception):
            logging.error(f"Не удалось уведомить {label} об отказе: {result}")


@dp.callback_query(F.data == "reject_skip", DriverRejectReason.waiting_for_reason)
async def reject_skip(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    lang = data.get("lang", "ru")
    await callback.message.edit_text(L[lang]["reject_processing"], reply_markup=None)
    await _do_reject(callback.message.chat.id, state, reason=None, photo_file_id=None)


@dp.message(DriverRejectReason.waiting_for_reason, F.photo)
async def reject_reason_photo(message: types.Message, state: FSMContext):
    reason = message.caption.strip() if message.caption else None
    photo_file_id = message.photo[-1].file_id
    await _do_reject(message.chat.id, state, reason=reason, photo_file_id=photo_file_id)


@dp.message(DriverRejectReason.waiting_for_reason, F.text)
async def reject_reason_text(message: types.Message, state: FSMContext):
    await _do_reject(message.chat.id, state, reason=message.text.strip(), photo_file_id=None)


@dp.callback_query(F.data.startswith("transit:"))
async def transit_order(callback: types.CallbackQuery):
    await callback.answer()
    driver_data = await _get_active_driver(callback.from_user.id)
    if not driver_data:
        await callback.message.answer(L["ru"]["access_denied_short"])
        return
    lang = _lang_from_driver_row(driver_data)
    order_id = callback.data.split(":", 1)[1]
    try:
        if not sheet:
            await callback.message.answer(L[lang]["db_unavailable"])
            return
        found = await asyncio.to_thread(_sync_find_order_by_id, order_id)
        client_chat_id = found[1][19] if found else ""
        ok = await asyncio.to_thread(
            _sync_update_status, order_id, "IN_TRANSIT", callback.from_user.id, {"TAKEN"}
        )
        if not ok:
            await callback.message.answer(L[lang]["access_denied_short"])
            return
        await asyncio.to_thread(sync_update_order_info_status, order_id, "IN_TRANSIT")
        b = InlineKeyboardBuilder()
        b.button(text=DELIVERED_BTN[lang], callback_data=f"done:{order_id}")
        await callback.message.edit_text(
            L[lang]["transit_msg"].format(id=order_id),
            reply_markup=b.as_markup(), parse_mode="Markdown"
        )
        if client_chat_id:
            await send_client_push(client_chat_id,
                ru=f"🚀 **Ваш заказ {order_id} уже в пути!**",
                tj=f"🚀 **Фармоиши шумо {order_id} дар роҳ аст!**")
    except Exception as e:
        logging.error(f"Ошибка transit (заказ {order_id}): {e}", exc_info=True)
        await callback.message.answer(L[lang]["transit_error"])


@dp.callback_query(F.data.startswith("done:"))
async def finish_order(callback: types.CallbackQuery):
    await callback.answer()
    driver_data = await _get_active_driver(callback.from_user.id)
    if not driver_data:
        await callback.message.answer(L["ru"]["access_denied_short"])
        return
    lang = _lang_from_driver_row(driver_data)
    order_id = callback.data.split(":", 1)[1]
    try:
        if not sheet:
            await callback.message.answer(L[lang]["db_unavailable"])
            return
        found = await asyncio.to_thread(_sync_find_order_by_id, order_id)
        client_chat_id = found[1][19] if found else ""
        ok = await asyncio.to_thread(
            _sync_update_status, order_id, "DELIVERED", callback.from_user.id, {"IN_TRANSIT"}
        )
        if not ok:
            await callback.message.answer(L[lang]["access_denied_short"])
            return
        await asyncio.to_thread(sync_update_order_info_status, order_id, "DELIVERED")
        delivered_at = datetime.now(DUSHANBE_TZ).strftime("%d.%m.%Y %H:%M")
        await asyncio.to_thread(sync_set_delivery_time, order_id, delivered_at)
        await callback.message.edit_text(
            L[lang]["delivered_msg"].format(id=order_id),
            reply_markup=None, parse_mode="Markdown"
        )
        if client_chat_id:
            await send_client_push(client_chat_id,
                ru=f"✅ **Ваш заказ {order_id} успешно доставлен!**\nСпасибо, что выбрали Mavsimi Rason!",
                tj=f"✅ **Фармоиши шумо {order_id} бомуваффақият расонида шуд!**\nМинатдорем, ки Mavsimi Rason-ро интихоб кардед!")
        # Пересобираем меню, чтобы кабинет (WebApp URL) сразу увидел новую доставку
        await callback.message.answer(
            L[lang]["menu_prompt"],
            reply_markup=await build_driver_main_menu(callback.from_user.id),
        )
    except Exception as e:
        logging.error(f"Ошибка done (заказ {order_id}): {e}", exc_info=True)
        await callback.message.answer(L[lang]["delivered_error"])


# ─── Поддержка курьеров (Topics) ─────────────────────────────────────────────
@dp.message(F.text.in_({SUPPORT_BTN["ru"], SUPPORT_BTN["tj"]}))
async def driver_support_start(message: types.Message, state: FSMContext):
    await _try_delete(message)
    driver_data = await _get_active_driver(message.from_user.id)
    if not driver_data:
        await message.answer(L["ru"]["access_denied"])
        return
    lang = _lang_from_driver_row(driver_data)
    if not SUPPORT_CHAT_ID:
        await message.answer(L[lang]["support_unavailable"])
        return
    await message.answer(
        L[lang]["support_prompt"],
        reply_markup=types.ReplyKeyboardRemove(),
        parse_mode="HTML"
    )
    await state.update_data(lang=lang)
    await state.set_state(DriverSupport.waiting_for_message)


@dp.message(DriverSupport.waiting_for_message, F.text)
async def driver_support_send(message: types.Message, state: FSMContext):
    driver_data = await asyncio.to_thread(_sync_get_driver, str(message.from_user.id))
    fio = driver_data[2] if driver_data and len(driver_data) > 2 else "Курьер"
    lang = _lang_from_driver_row(driver_data)
    chat_id_str = str(message.from_user.id)
    try:
        topic_id_str = await asyncio.to_thread(_sync_get_driver_support_topic, chat_id_str)
        if topic_id_str:
            topic_id = int(topic_id_str)
        else:
            topic = await bot.create_forum_topic(
                chat_id=int(SUPPORT_CHAT_ID),
                name=f"🚗 Курьер: {fio}"
            )
            topic_id = topic.message_thread_id
            await asyncio.to_thread(_sync_save_driver_support_topic, chat_id_str, topic_id)

        await bot.send_message(
            chat_id=int(SUPPORT_CHAT_ID),
            message_thread_id=topic_id,
            text=f"🚗 <b>Курьер:</b> {html.escape(message.text)}",
            parse_mode="HTML"
        )
        back_kb = ReplyKeyboardBuilder()
        back_kb.button(text=BACK_BTN[lang])
        await state.update_data(fio=fio, topic_id=topic_id, lang=lang)
        await state.set_state(DriverSupport.chatting)
        await message.answer(
            L[lang]["support_sent"],
            reply_markup=back_kb.as_markup(resize_keyboard=True),
        )
    except Exception as e:
        await state.clear()
        logging.error(f"Ошибка поддержки курьера (SUPPORT_CHAT_ID={SUPPORT_CHAT_ID}): {type(e).__name__}: {e}")
        await message.answer(L[lang]["support_error"], reply_markup=await build_driver_main_menu(message.from_user.id))


@dp.message(DriverSupport.chatting, F.text)
async def driver_support_continue(message: types.Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("lang", "ru")
    topic_id = data.get("topic_id")
    if not topic_id:
        await state.clear()
        await message.answer(L[lang]["support_session_expired"], reply_markup=await build_driver_main_menu(message.from_user.id))
        return
    try:
        await bot.send_message(
            chat_id=int(SUPPORT_CHAT_ID),
            message_thread_id=topic_id,
            text=f"🚗 <b>Курьер:</b> {html.escape(message.text)}",
            parse_mode="HTML"
        )
    except Exception as e:
        logging.error(f"Ошибка отправки сообщения курьера в топик: {type(e).__name__}: {e}")
        await message.answer(L[lang]["support_send_failed"])


@dp.message(DriverSupport.waiting_for_message)
@dp.message(DriverSupport.chatting)
async def driver_support_non_text(message: types.Message, state: FSMContext):
    """Фото/стикер/голос в поддержке: раньше падало на html.escape(None) и курьер
    не получал вообще ничего в ответ."""
    data = await state.get_data()
    await message.answer(L[data.get("lang", "ru")]["support_text_only"])


@dp.message(F.chat.type.in_({"group", "supergroup"}))
async def driver_support_group_message(message: types.Message):
    if not SUPPORT_CHAT_ID or str(message.chat.id) != str(SUPPORT_CHAT_ID):
        return
    if not message.message_thread_id:
        return
    if not message.from_user or message.from_user.is_bot:
        return
    if not message.text:
        return
    driver_telegram_id = await asyncio.to_thread(
        _sync_get_driver_by_topic, str(message.message_thread_id)
    )
    if not driver_telegram_id:
        return
    driver_data = await asyncio.to_thread(_sync_get_driver, str(driver_telegram_id))
    lang = _lang_from_driver_row(driver_data)
    try:
        await bot.send_message(
            chat_id=int(driver_telegram_id),
            text=L[lang]["support_reply_header"].format(text=html.escape(message.text)),
            parse_mode="HTML"
        )
    except Exception as e:
        logging.error(f"Ошибка отправки ответа курьеру {driver_telegram_id}: {type(e).__name__}: {e}")


# ─── Автоуведомление курьеров о новых заказах ────────────────────────────────
_notified_order_ids: set[str] = set()


async def _broadcast_new_free_orders():
    while True:
        try:
            free_jobs = await asyncio.to_thread(_sync_get_free_orders)
            current_ids = {o["id"] for o in free_jobs}
            new_jobs = [o for o in free_jobs if o["id"] not in _notified_order_ids]

            if new_jobs:
                drivers = await asyncio.to_thread(_sync_get_all_active_drivers)
                if not drivers:
                    logging.info("Новые заказы есть, но активных курьеров нет — повторим на следующей проверке.")
                for o in (new_jobs if drivers else []):
                    for d in drivers:
                        lang = d.get("lang", "ru")
                        card = _render_job_card(o, lang, header=L[lang]["new_job_header"])
                        b = InlineKeyboardBuilder()
                        b.button(text=TAKE_JOB_BTN[lang], callback_data=f"take:{o['id']}")
                        try:
                            sent = await bot.send_message(
                                d["telegram_id"], card,
                                reply_markup=b.as_markup(), parse_mode="Markdown"
                            )
                            _job_message_refs.setdefault(o['id'], []).append((int(d["telegram_id"]), sent.message_id))
                        except Exception as e:
                            logging.warning(f"Не удалось отправить пуш курьеру {d['telegram_id']}: {e}")
                        # Telegram режет отправку быстрее ~30 msg/s: при большой базе
                        # курьеров рассылка без паузы упирается в flood control.
                        await asyncio.sleep(BROADCAST_DELAY_SECONDS)

                    _notified_order_ids.add(o["id"])

            # заказ забрали/отменили — освобождаем id, чтобы он мог снова попасть в рассылку при повторном появлении
            _notified_order_ids.intersection_update(current_ids)
            # и чистим трекинг разосланных карточек по тем же id — иначе _job_message_refs
            # растёт без предела (заказы уходят из свободных, а записи о карточках висят вечно)
            for stale_id in [oid for oid in _job_message_refs if oid not in current_ids]:
                _job_message_refs.pop(stale_id, None)
        except Exception:
            logging.error(f"Сбой автоуведомления о новых заказах: {traceback.format_exc()}")

        await asyncio.sleep(NEW_ORDERS_POLL_SECONDS)


@dp.startup()
async def _on_driver_bot_startup(**kwargs):
    asyncio.create_task(_broadcast_new_free_orders())