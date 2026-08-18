import asyncio
import base64
import html
import json
import logging
import os
import time
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from aiogram import types, F
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.utils.keyboard import InlineKeyboardBuilder

from config import (
    manager_bot as bot,
    manager_dp as dp,
    driver_bot as driver_bot_instance,
    client_bot,
    sheet,
    drivers_sheet,
    get_manager_chat_ids,
    sync_update_order_info_status,
    md_escape,
)
from driver_bot import (
    _pad_row,
    _sync_find_order_by_id,
    _sync_reassign_order,
    _sync_get_all_active_drivers,
    _async_get_admin_dashboard_data,
    _sync_get_driver,
    _sync_get_driver_deliveries,
    _sync_approve_name_change,
    _sync_reject_name_change,
    _lang_from_driver_row,
    _clear_status_message,
    L as DRIVER_L,
    _week_label,
    generate_excel_report,
    send_client_push,
)

import db

ADMIN_PANEL_URL     = os.getenv("ADMIN_PANEL_URL", "")
DASHBOARD_URL       = os.getenv("DASHBOARD_URL", "")
DEFAULT_DRIVER_RATE = float(os.getenv("DEFAULT_DRIVER_RATE", "18.0"))

# Ручной переключатель статуса в панели пишет значение прямо в таблицу —
# без whitelist любая опечатка/подмена в WebApp порождала мусорный статус,
# который потом никто не умеет обработать.
ALLOWED_ORDER_STATUSES = {
    "NEW", "READY_FOR_DRIVERS", "TAKEN", "IN_TRANSIT", "DELIVERED", "CANCELLED",
}

CANCEL_REASONS = [
    "Не смогли согласовать детали",
    "Адрес недоступен",
    "Нет курьеров в этой зоне",
    "Дублирующий заказ",
]


class ManagerCancelOrder(StatesGroup):
    waiting_for_reason = State()


_managers_cache: dict = {"ids": set(), "ts": 0.0}
_MANAGERS_CACHE_TTL = 60  # секунд, чтобы не дёргать Google Sheets на каждое сообщение


async def _is_manager(chat_id) -> bool:
    if time.time() - _managers_cache["ts"] > _MANAGERS_CACHE_TTL:
        try:
            _managers_cache["ids"] = set(await asyncio.to_thread(get_manager_chat_ids))
            _managers_cache["ts"] = time.time()
        except Exception as e:
            logging.error(f"Ошибка обновления списка менеджеров: {e}")
    return str(chat_id) in _managers_cache["ids"]


# ─── Google Sheets: курьеры ──────────────────────────────────────────────────

def _sync_approve_driver(telegram_id: str) -> str | None:
    if not drivers_sheet:
        return None
    try:
        cell = drivers_sheet.find(str(telegram_id), in_column=4)
        if not cell:
            return None
        row = drivers_sheet.row_values(cell.row)
        if row[0].upper().strip() != "PENDING":
            return None  # уже обработан (повторный тап/дублирующий callback) — не шлём уведомление снова
        drivers_sheet.update_cell(cell.row, 1, "ACTIVE")
        db.mark_driver_status(telegram_id, "ACTIVE")
        return row[2] if len(row) > 2 else "Курьер"
    except Exception as e:
        logging.error(f"Ошибка одобрения курьера {telegram_id}: {e}")
        return None


def _sync_reject_driver(telegram_id: str) -> str | None:
    if not drivers_sheet:
        return None
    try:
        cell = drivers_sheet.find(str(telegram_id), in_column=4)
        if not cell:
            return None
        row = drivers_sheet.row_values(cell.row)
        if row[0].upper().strip() != "PENDING":
            return None  # уже обработан (повторный тап/дублирующий callback) — не шлём уведомление снова
        drivers_sheet.update_cell(cell.row, 1, "REJECTED")
        db.mark_driver_status(telegram_id, "REJECTED")
        return row[2] if len(row) > 2 else "Курьер"
    except Exception as e:
        logging.error(f"Ошибка отклонения курьера {telegram_id}: {e}")
        return None


def _sync_set_driver_access(telegram_id: str, new_status: str, allowed_from: tuple) -> tuple[bool, str, str]:
    """Блокировка/разблокировка курьера. Returns (успех, ФИО, ошибка).

    Раньше отзыв доступа делался правкой листа «Водители» руками — код об этом
    не знал, и зеркало продолжало отдавать ACTIVE до следующего снапшота.
    Здесь статус меняется вместе с write-through, поэтому доступ пропадает
    сразу же (гейт `_get_active_driver` читает то же зеркало)."""
    if not drivers_sheet:
        return False, "", "база недоступна"
    try:
        cell = drivers_sheet.find(str(telegram_id), in_column=4)
        if not cell:
            return False, "", "курьер не найден"
        row = drivers_sheet.row_values(cell.row)
        current = row[0].upper().strip() if row else ""
        if current not in allowed_from:
            return False, "", f"статус: {current or '—'}"
        drivers_sheet.update_cell(cell.row, 1, new_status)
        db.mark_driver_status(telegram_id, new_status)
        return True, (row[2] if len(row) > 2 else "Курьер"), ""
    except Exception as e:
        logging.error(f"Ошибка смены доступа курьера {telegram_id} на {new_status}: {e}")
        return False, "", str(e)


# ─── Google Sheets: заказы ───────────────────────────────────────────────────

def _sync_set_order_ready(order_id: str) -> tuple[bool, str, str]:
    """NEW → READY_FOR_DRIVERS. Returns (success, client_chat_id, error_detail)."""
    if not sheet:
        return False, "", "база недоступна"
    try:
        cell = sheet.find(str(order_id), in_column=2)
        if not cell:
            return False, "", "не найден"
        row = _pad_row(sheet.row_values(cell.row))
        status = row[0].upper().strip()
        if status != "NEW":
            return False, "", f"статус: {status}"
        sheet.update_cell(cell.row, 1, "READY_FOR_DRIVERS")
        sync_update_order_info_status(order_id, "READY_FOR_DRIVERS")
        # upsert, а не UPDATE: заказ мог быть создан после последнего снапшота,
        # и тогда в зеркале его ещё нет — биржа не увидела бы его до снапшота.
        row[0] = "READY_FOR_DRIVERS"
        db.upsert_order_from_row(row)
        return True, row[19], ""
    except Exception as e:
        logging.error(f"Ошибка перевода заказа {order_id} в READY: {e}")
        return False, "", str(e)


def _sync_change_order_status(order_id: str, new_status: str) -> tuple[bool, str, str, str]:
    """Меняет статус заказа. Returns (success, client_chat_id, courier_id, error)."""
    if not sheet:
        return False, "", "", "база недоступна"
    if new_status not in ALLOWED_ORDER_STATUSES:
        return False, "", "", f"недопустимый статус: {new_status}"
    try:
        cell = sheet.find(str(order_id), in_column=2)
        if not cell:
            return False, "", "", "не найден"
        row = _pad_row(sheet.row_values(cell.row))
        sheet.update_cell(cell.row, 1, new_status)
        sync_update_order_info_status(order_id, new_status)
        row[0] = new_status
        db.upsert_order_from_row(row)
        return True, row[19], row[20], ""
    except Exception as e:
        logging.error(f"Ошибка смены статуса заказа {order_id}: {e}")
        return False, "", "", str(e)


def _sync_cancel_order(order_id: str) -> tuple[bool, str, str]:
    """NEW → CANCELLED. Returns (success, client_chat_id, error_detail)."""
    if not sheet:
        return False, "", "база недоступна"
    try:
        cell = sheet.find(str(order_id), in_column=2)
        if not cell:
            return False, "", "не найден"
        row = _pad_row(sheet.row_values(cell.row))
        status = row[0].upper().strip()
        if status != "NEW":
            return False, "", f"статус: {status}"
        sheet.update_cell(cell.row, 1, "CANCELLED")
        sync_update_order_info_status(order_id, "CANCELLED")
        db.mark_order_status(order_id, "CANCELLED")
        return True, row[19], ""
    except Exception as e:
        logging.error(f"Ошибка отмены заказа {order_id}: {e}")
        return False, "", str(e)


# ─── Google Sheets: сводные данные по всем курьерам ─────────────────────────

def _sync_get_all_couriers_deliveries(date_from: datetime, date_to: datetime) -> dict:
    """Returns {telegram_id: {total: N, delivered: M}} for the given date range."""
    if not sheet:
        return {}
    try:
        result = {}
        for idx, row in enumerate(sheet.get_all_values()):
            if idx == 0:
                continue
            row = _pad_row(row)
            courier_id = str(row[20]).strip()
            if not courier_id:
                continue
            date_cell = row[2].strip()
            try:
                dt = datetime.strptime(date_cell[:16], "%d.%m.%Y %H:%M")
            except ValueError:
                continue
            if not (date_from <= dt <= date_to):
                continue
            if courier_id not in result:
                result[courier_id] = {"total": 0, "delivered": 0}
            result[courier_id]["total"] += 1
            if row[0].upper().strip() == "DELIVERED":
                result[courier_id]["delivered"] += 1
        return result
    except Exception as e:
        logging.error(f"Ошибка чтения доставок всех курьеров: {e}")
        return {}


def _sync_get_all_drivers_rates() -> dict:
    """Returns {telegram_id: rate} for all ACTIVE drivers."""
    if not drivers_sheet:
        return {}
    try:
        result = {}
        for idx, row in enumerate(drivers_sheet.get_all_values()):
            if idx == 0 or len(row) < 5:
                continue
            if row[0].upper().strip() != "ACTIVE":
                continue
            tid = str(row[3]).strip()
            try:
                rate = float(row[4]) if row[4] else DEFAULT_DRIVER_RATE
            except (ValueError, TypeError):
                rate = DEFAULT_DRIVER_RATE
            result[tid] = rate
        return result
    except Exception as e:
        logging.error(f"Ошибка чтения ставок курьеров: {e}")
        return {}


# ─── Excel: сводный отчёт на всех курьеров ───────────────────────────────────

def generate_summary_excel(couriers_stats: list[dict], period_label: str) -> BytesIO:
    """
    couriers_stats: [{fio, rate, total, delivered}]
    Generates a styled summary Excel — one row per courier + totals.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводка"

    BRAND   = "FFEA580C"
    GREEN   = "FF1F9D4D"
    WHITE  = "FFFFFFFF"
    GRAY   = "FFF4F4F5"
    L_BRAND = "FFFDECE0"

    thin = Side(border_style="thin", color="FFD0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sc(row, col, val, bold=False, bg=None, color="FF1C1C1E", align="center", size=11):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=color, size=size, name="Calibri")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        return c

    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value="MAVSIMI RASON — Сводный отчёт")
    t.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    t.fill = PatternFill("solid", fgColor=BRAND)
    t.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    t2 = ws.cell(row=2, column=1, value=f"Период: {period_label}")
    t2.font = Font(color="FF555555", size=11, name="Calibri")
    t2.fill = PatternFill("solid", fgColor=GRAY)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["№", "ФИО курьера", "Ставка (TJS)", "Заказов взято", "Доставлено", "К выплате (TJS)"]
    for col, h in enumerate(headers, 1):
        sc(3, col, h, bold=True, bg=BRAND, color=WHITE)

    col_widths = [4, 26, 13, 14, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    total_delivered = 0
    total_earnings  = 0.0

    for i, cs in enumerate(couriers_stats, 1):
        r  = i + 3
        bg = L_BRAND if i % 2 == 0 else WHITE
        delivered = cs["delivered"]
        rate      = cs["rate"]
        earnings  = delivered * rate
        total_delivered += delivered
        total_earnings  += earnings
        sc(r, 1, i,             bg=bg)
        sc(r, 2, cs["fio"],     bg=bg, align="left")
        sc(r, 3, rate,          bg=bg)
        sc(r, 4, cs["total"],   bg=bg)
        sc(r, 5, delivered,     bg=bg, bold=True)
        ec = sc(r, 6, earnings, bg=bg, bold=True, color="FFEA580C")
        ec.number_format = '0.00 "TJS"'
        ws.row_dimensions[r].height = 18

    last = len(couriers_stats) + 4
    ws.merge_cells(f"A{last}:E{last}")
    lbl = ws.cell(row=last, column=1, value="ИТОГО К ВЫПЛАТЕ:")
    lbl.font = Font(bold=True, size=12, name="Calibri")
    lbl.fill = PatternFill("solid", fgColor=GRAY)
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.border = border
    tot = ws.cell(row=last, column=6, value=total_earnings)
    tot.font = Font(bold=True, color=GREEN, size=13, name="Calibri")
    tot.fill = PatternFill("solid", fgColor=GRAY)
    tot.alignment = Alignment(horizontal="center", vertical="center")
    tot.number_format = '0.00 "TJS"'
    tot.border = border
    ws.row_dimensions[last].height = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Панель управления ───────────────────────────────────────────────────────

def _build_panel_message(data: dict) -> tuple[str, str | None]:
    active_cnt = len(data["orders"])
    free_cnt   = len(data["free"])
    # Заблокированные приходят в списке (их надо показать в панели), но в
    # счётчик «курьеров в работе» не входят.
    working    = [c for c in data["couriers"] if c.get("status") != "BLOCKED"]
    busy_cnt   = sum(1 for c in working if c["busy"])
    total_cnt  = len(working)
    text = (
        f"🎛 <b>Панель управления</b>\n\n"
        f"📦 Активных заказов: <b>{active_cnt}</b>\n"
        f"🕳️ Свободных заказов: <b>{free_cnt}</b>\n"
        f"🚗 Курьеров в работе: <b>{busy_cnt}/{total_cnt}</b>"
    )
    webapp_url = None
    if ADMIN_PANEL_URL:
        raw = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
        b64 = base64.urlsafe_b64encode(raw.encode()).decode().rstrip("=")
        webapp_url = f"{ADMIN_PANEL_URL}?d={b64}"
    return text, webapp_url


async def _build_dashboard_url() -> str | None:
    """Ссылка на дашборд статистики со свежим срезом из базы-зеркала (SQLite).
    Отдельная страница, не связана с личной панелью менеджера."""
    if not DASHBOARD_URL:
        return None
    try:
        payload = await asyncio.to_thread(db.dashboard_payload)
        raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        b64 = base64.urlsafe_b64encode(raw.encode()).decode().rstrip("=")
        return f"{DASHBOARD_URL}?d={b64}"
    except Exception as e:
        logging.error(f"Не удалось собрать дашборд статистики: {e}")
        return None


@dp.message(CommandStart())
async def cmd_start_manager(message: types.Message):
    if not await _is_manager(message.chat.id):
        await message.answer("⛔ Доступ запрещён.")
        return
    await _send_panel(message.chat.id, bot)


async def _send_panel(chat_id: int, bot_instance):
    data = await _async_get_admin_dashboard_data()
    text, webapp_url = _build_panel_message(data)
    dashboard_url = await _build_dashboard_url()
    rows = []
    if webapp_url:
        rows.append([KeyboardButton(text="🎛 Открыть панель", web_app=types.WebAppInfo(url=webapp_url))])
    if dashboard_url:
        rows.append([KeyboardButton(text="📊 Статистика", web_app=types.WebAppInfo(url=dashboard_url))])
    rows.append([KeyboardButton(text="🔄 Обновить")])
    reply_kb = ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)
    await bot_instance.send_message(chat_id, text, reply_markup=reply_kb, parse_mode="HTML")



@dp.message(F.text == "🔄 Обновить")
async def panel_refresh_text(message: types.Message):
    if not await _is_manager(message.chat.id):
        return
    await _send_panel(message.chat.id, bot)




# ─── Принятие / отмена нового заказа (push от клиентского бота) ─────────────

@dp.callback_query(F.data.startswith("oa:"))
async def order_accept(callback: types.CallbackQuery):
    await callback.answer()
    if not await _is_manager(callback.from_user.id):
        return
    order_id = callback.data.split(":", 1)[1]
    success, client_chat_id, err = await asyncio.to_thread(_sync_set_order_ready, order_id)
    if not success:
        await callback.message.edit_text(
            callback.message.html_text + f"\n\n❌ Не удалось принять — {html.escape(err)}.",
            reply_markup=None, parse_mode="HTML"
        )
        return
    await callback.message.edit_text(
        callback.message.html_text + "\n\n✅ <b>Принят и передан на биржу!</b>",
        reply_markup=None,
        parse_mode="HTML"
    )
    if client_chat_id:
        await send_client_push(
            client_chat_id,
            f"✅ Ваш заказ *{order_id}* принят!\nОжидайте назначения курьера."
        )


@dp.callback_query(F.data.startswith("oc:"))
async def order_cancel_menu(callback: types.CallbackQuery):
    await callback.answer()
    if not await _is_manager(callback.from_user.id):
        return
    order_id = callback.data.split(":", 1)[1]
    b = InlineKeyboardBuilder()
    for i, reason in enumerate(CANCEL_REASONS):
        b.button(text=reason, callback_data=f"ocr:{order_id}:{i}")
    b.button(text="✏️ Своя причина", callback_data=f"ocx:{order_id}")
    b.adjust(1)
    await callback.message.edit_reply_markup(reply_markup=b.as_markup())


@dp.callback_query(F.data.startswith("ocr:"))
async def order_cancel_reason(callback: types.CallbackQuery):
    await callback.answer()
    if not await _is_manager(callback.from_user.id):
        return
    _, order_id, idx_str = callback.data.split(":", 2)
    reason = CANCEL_REASONS[int(idx_str)]
    success, client_chat_id, err = await asyncio.to_thread(_sync_cancel_order, order_id)
    if not success:
        await callback.message.edit_text(
            callback.message.html_text + f"\n\n❌ Не удалось отменить — {err}.",
            reply_markup=None, parse_mode="HTML"
        )
        return
    await callback.message.edit_text(
        callback.message.html_text + f"\n\n❌ <b>Отменён.</b> Причина: {reason}",
        reply_markup=None, parse_mode="HTML"
    )
    if client_chat_id:
        await send_client_push(
            client_chat_id,
            f"❌ К сожалению, ваш заказ *{order_id}* был отменён.\n📝 Причина: {reason}\n\nПожалуйста, свяжитесь с поддержкой если возникли вопросы."
        )


@dp.callback_query(F.data.startswith("ocx:"))
async def order_cancel_custom_start(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    if not await _is_manager(callback.from_user.id):
        return
    order_id = callback.data.split(":", 1)[1]
    await state.set_state(ManagerCancelOrder.waiting_for_reason)
    await state.update_data(order_id=order_id, msg_id=callback.message.message_id)
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(f"✏️ Введите причину отмены заказа <code>{order_id}</code>:", parse_mode="HTML")


@dp.message(ManagerCancelOrder.waiting_for_reason, F.text)
async def order_cancel_custom_reason(message: types.Message, state: FSMContext):
    if not await _is_manager(message.chat.id):
        return
    data = await state.get_data()
    order_id = data.get("order_id", "")
    reason = message.text.strip()
    await state.clear()

    success, client_chat_id, err = await asyncio.to_thread(_sync_cancel_order, order_id)
    if not success:
        await message.answer(f"❌ Не удалось отменить заказ {order_id} — {err}.")
        return
    await message.answer(f"❌ Заказ <code>{order_id}</code> отменён.\nПричина: {html.escape(reason)}", parse_mode="HTML")
    if client_chat_id:
        await send_client_push(
            client_chat_id,
            f"❌ К сожалению, ваш заказ *{order_id}* был отменён.\n📝 Причина: {md_escape(reason)}\n\nПожалуйста, свяжитесь с поддержкой если возникли вопросы."
        )


# ─── WebApp от панели менеджера ──────────────────────────────────────────────

@dp.message(F.web_app_data)
async def handle_webapp(message: types.Message):
    if not await _is_manager(message.chat.id):
        return
    try:
        data = json.loads(message.web_app_data.data)
    except Exception:
        return

    action = data.get("action")

    # ── Экспорт реестра для бухгалтерии (из дашборда статистики) ─────────────
    if action == "export_accounting":
        period = data.get("period", "month")
        try:
            buf, fname, caption = await asyncio.to_thread(db.build_accounting_excel, period)
        except Exception as e:
            logging.error(f"Ошибка экспорта бухгалтерии ({period}): {e}")
            await message.answer("❌ Не удалось сформировать реестр. Попробуйте позже.")
            return
        await message.answer_document(
            types.BufferedInputFile(buf.read(), filename=fname),
            caption=caption,
        )
        return

    # ── Заблокировать / разблокировать курьера ──────────────────────────────
    if action in ("block_driver", "unblock_driver"):
        tid = str(data.get("courier_tid") or "").strip()
        if not tid.isdigit():
            return
        blocking    = action == "block_driver"
        new_status  = "BLOCKED" if blocking else "ACTIVE"
        # Блокируем только активного, разблокируем только заблокированного:
        # иначе панель со старым снимком могла бы «разблокировать» курьера,
        # которого на самом деле отклонили при регистрации (REJECTED).
        allowed_from = ("ACTIVE",) if blocking else ("BLOCKED",)

        ok, fio, err = await asyncio.to_thread(
            _sync_set_driver_access, tid, new_status, allowed_from
        )
        if not ok:
            await message.answer(f"❌ Не удалось изменить доступ курьера. {err}")
            return

        note = ""
        if blocking:
            stuck = await asyncio.to_thread(db.count_active_orders_for_courier, tid)
            if stuck:
                note = (f"\n\n⚠️ На нём висит незакрытых заказов: <b>{stuck}</b>. "
                        f"Переназначьте их через карточку заказа.")
        await message.answer(
            (f"⛔ Курьер <b>{html.escape(fio)}</b> заблокирован. Доступ к заказам закрыт.{note}"
             if blocking else
             f"✅ Курьер <b>{html.escape(fio)}</b> разблокирован."),
            parse_mode="HTML"
        )

        try:
            driver_data = await asyncio.to_thread(_sync_get_driver, tid)
            lang = _lang_from_driver_row(driver_data)
            await driver_bot_instance.send_message(
                chat_id=int(tid),
                text=DRIVER_L[lang]["access_revoked" if blocking else "access_restored"],
            )
        except Exception as e:
            logging.error(f"Не удалось уведомить курьера {tid} о смене доступа: {e}")
        return

    # ── Сменить статус активного заказа ─────────────────────────────────────
    if action == "change_status":
        order_id       = data.get("order_id")
        new_status     = data.get("new_status")
        new_status_lbl = data.get("new_status_label", new_status)
        if not order_id or not new_status:
            return
        success, client_chat_id, courier_id, err = await asyncio.to_thread(
            _sync_change_order_status, order_id, new_status
        )
        if not success:
            await message.answer(f"❌ Не удалось сменить статус — {err}.")
            return
        await message.answer(
            f"✅ Заказ <b>{order_id}</b> → <b>{html.escape(str(new_status_lbl))}</b>",
            parse_mode="HTML"
        )
        if courier_id:
            try:
                await driver_bot_instance.send_message(
                    chat_id=int(courier_id),
                    text=f"📋 <b>Статус заказа {order_id} изменён менеджером:</b> {html.escape(str(new_status_lbl))}",
                    parse_mode="HTML"
                )
            except Exception as e:
                logging.error(f"Не удалось уведомить курьера {courier_id}: {e}")
        if client_chat_id:
            await send_client_push(
                client_chat_id,
                f"📦 Статус вашего заказа *{order_id}* обновлён: *{md_escape(new_status_lbl)}*"
            )
        await _send_panel(message.chat.id, bot)
        return

    # ── Отменить активный заказ (с курьером) ────────────────────────────────
    if action == "cancel_active":
        order_id   = data.get("order_id")
        reason     = data.get("reason", "Отменён менеджером")
        if not order_id:
            return
        success, client_chat_id, courier_id, err = await asyncio.to_thread(
            _sync_change_order_status, order_id, "CANCELLED"
        )
        if not success:
            await message.answer(f"❌ Не удалось отменить — {err}.")
            return
        await message.answer(
            f"❌ Заказ <b>{order_id}</b> отменён.\nПричина: {html.escape(str(reason))}",
            parse_mode="HTML"
        )
        if courier_id:
            try:
                await driver_bot_instance.send_message(
                    chat_id=int(courier_id),
                    text=f"⚠️ <b>Заказ {order_id} отменён менеджером.</b>\nПричина: {html.escape(str(reason))}",
                    parse_mode="HTML"
                )
            except Exception as e:
                logging.error(f"Не удалось уведомить курьера {courier_id}: {e}")
        if client_chat_id:
            await send_client_push(
                client_chat_id,
                f"❌ Ваш заказ *{order_id}* был отменён.\n📝 Причина: {md_escape(reason)}"
            )
        await _send_panel(message.chat.id, bot)
        return

    # ── Принять NEW-заказ → READY_FOR_DRIVERS ───────────────────────────────
    if action == "set_ready":
        order_id = data.get("order_id")
        if not order_id:
            return
        success, client_chat_id, err = await asyncio.to_thread(_sync_set_order_ready, order_id)
        if not success:
            await message.answer(f"❌ Не удалось принять заказ — {err}.")
            return
        await message.answer(
            f"✅ Заказ <b>{order_id}</b> принят и передан на биржу курьеров!",
            parse_mode="HTML"
        )
        if client_chat_id:
            await send_client_push(
                client_chat_id,
                f"📦 Ваш заказ *{order_id}* принят и передан в доставку!\nОжидайте назначения курьера."
            )
        await _send_panel(message.chat.id, bot)
        return

    # ── Переназначение — выбор курьера внутри WebApp ────────────────────────
    if action == "reassign_confirm":
        # order_row из панели больше не используется: номер строки в таблице
        # непостоянен, работаем строго по ID заказа.
        order_id     = data.get("order_id")
        courier_tid  = str(data.get("courier_tid", ""))
        courier_name = data.get("courier_name", "")
        if not order_id or not courier_tid:
            return
        success, old_courier_id, confirmed_order_id = await asyncio.to_thread(
            _sync_reassign_order, order_id, courier_name, courier_tid
        )
        if not success:
            await message.answer("❌ Не удалось переназначить. Проверьте статус заказа.")
            return
        found = await asyncio.to_thread(_sync_find_order_by_id, confirmed_order_id)
        order_row_vals = found[1] if found else _pad_row([])
        client_chat_id = order_row_vals[19]
        city_from      = order_row_vals[5]
        city_to        = order_row_vals[7]
        await message.answer(
            f"✅ Заказ <b>{confirmed_order_id}</b> переназначен → <b>{html.escape(str(courier_name))}</b>",
            parse_mode="HTML"
        )
        if old_courier_id and old_courier_id != courier_tid:
            try:
                await driver_bot_instance.send_message(
                    chat_id=int(old_courier_id),
                    text=f"⚠️ <b>Ваш заказ {confirmed_order_id} переназначен другому курьеру.</b>",
                    parse_mode="HTML"
                )
            except Exception as e:
                logging.error(f"Не удалось уведомить старого курьера: {e}")
        b = InlineKeyboardBuilder()
        b.button(text="🚗 В путь", callback_data=f"transit:{confirmed_order_id}")
        try:
            await driver_bot_instance.send_message(
                chat_id=int(courier_tid),
                text=(
                    f"📦 <b>Вам назначен заказ {confirmed_order_id}!</b>\n"
                    f"📍 {html.escape(str(city_from))} → {html.escape(str(city_to))}\n\n"
                    f"Нажмите кнопку, когда выедете:"
                ),
                reply_markup=b.as_markup(),
                parse_mode="HTML"
            )
        except Exception as e:
            logging.error(f"Не удалось уведомить нового курьера {courier_tid}: {e}")
        if client_chat_id:
            await send_client_push(
                client_chat_id,
                f"🔄 Ваш заказ *{confirmed_order_id}* передан новому курьеру: *{md_escape(courier_name)}*."
            )
        await _send_panel(message.chat.id, bot)
        return

    # ── Отчёт на одного курьера ──────────────────────────────────────────────
    if action == "manager_report":
        courier_tid    = str(data.get("courier_tid", ""))
        courier_name   = data.get("courier_name", "Курьер")
        week_start_str = data.get("week_start", "")
        if not courier_tid or not week_start_str:
            return
        try:
            week_start = datetime.strptime(week_start_str, "%Y-%m-%d")
        except ValueError:
            await message.answer("❌ Некорректный формат даты.")
            return
        week_end     = week_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
        period_label = _week_label(week_start, week_end)

        driver_data = await asyncio.to_thread(_sync_get_driver, courier_tid)
        try:
            rate = float(driver_data[4]) if driver_data and len(driver_data) > 4 and driver_data[4] else DEFAULT_DRIVER_RATE
        except (ValueError, TypeError):
            rate = DEFAULT_DRIVER_RATE

        wait = await message.answer(f"⏳ Формирую отчёт для {courier_name}...")
        deliveries = await asyncio.to_thread(_sync_get_driver_deliveries, courier_tid, week_start, week_end)

        if not deliveries:
            await wait.delete()
            await message.answer(f"📭 За {period_label} у курьера <b>{html.escape(courier_name)}</b> доставок не найдено.", parse_mode="HTML")
            return

        excel_buf = await asyncio.to_thread(generate_excel_report, courier_name, rate, deliveries, period_label)
        delivered_count = sum(1 for d in deliveries if d["s"] == "DELIVERED")
        await wait.delete()
        await message.answer_document(
            types.BufferedInputFile(excel_buf.read(), filename=f"report_{courier_tid}_{week_start_str}.xlsx"),
            caption=(
                f"📄 <b>Отчёт: {period_label}</b>\n"
                f"👤 {html.escape(courier_name)}\n"
                f"✅ Доставлено: {delivered_count}\n"
                f"💰 К выплате: {delivered_count * rate:.2f} TJS"
            ),
            parse_mode="HTML"
        )
        return

    # ── Сводный отчёт на всех курьеров ───────────────────────────────────────
    if action == "manager_report_all":
        week_start_str = data.get("week_start", "")
        if not week_start_str:
            return
        try:
            week_start = datetime.strptime(week_start_str, "%Y-%m-%d")
        except ValueError:
            await message.answer("❌ Некорректный формат даты.")
            return
        week_end     = week_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
        period_label = _week_label(week_start, week_end)

        active_couriers = await asyncio.to_thread(_sync_get_all_active_drivers)
        if not active_couriers:
            await message.answer("❌ Нет активных курьеров.")
            return

        wait = await message.answer(f"⏳ Формирую сводный отчёт за {period_label}...")

        deliveries_map, rates_map = await asyncio.gather(
            asyncio.to_thread(_sync_get_all_couriers_deliveries, week_start, week_end),
            asyncio.to_thread(_sync_get_all_drivers_rates),
        )

        couriers_stats = []
        for c in active_couriers:
            rate  = rates_map.get(c["telegram_id"], DEFAULT_DRIVER_RATE)
            stats = deliveries_map.get(c["telegram_id"], {"total": 0, "delivered": 0})
            couriers_stats.append({
                "fio":       c["fio"],
                "rate":      rate,
                "total":     stats["total"],
                "delivered": stats["delivered"],
            })

        excel_buf   = await asyncio.to_thread(generate_summary_excel, couriers_stats, period_label)
        total_earn  = sum(c["delivered"] * c["rate"] for c in couriers_stats)
        total_deliv = sum(c["delivered"] for c in couriers_stats)
        await wait.delete()
        await message.answer_document(
            types.BufferedInputFile(excel_buf.read(), filename=f"summary_{week_start_str}.xlsx"),
            caption=(
                f"📊 <b>Сводный отчёт: {period_label}</b>\n"
                f"👥 Курьеров: {len(couriers_stats)}\n"
                f"✅ Доставлено: {total_deliv}\n"
                f"💰 Итого к выплате: {total_earn:.2f} TJS"
            ),
            parse_mode="HTML"
        )
        return


# ─── Одобрение / отклонение курьеров ────────────────────────────────────────

@dp.callback_query(F.data.startswith("approve_driver:"))
async def approve_driver(callback: types.CallbackQuery):
    await callback.answer()
    if not await _is_manager(callback.from_user.id):
        return
    telegram_id = callback.data.split(":", 1)[1]
    fio = await asyncio.to_thread(_sync_approve_driver, telegram_id)
    if not fio:
        await callback.message.edit_text("❌ Курьер не найден или уже обработан.", reply_markup=None)
        return
    await callback.message.edit_text(
        f"✅ Курьер <b>{html.escape(fio)}</b> одобрен и активирован.",
        reply_markup=None, parse_mode="HTML"
    )
    try:
        driver_data = await asyncio.to_thread(_sync_get_driver, telegram_id)
        lang = _lang_from_driver_row(driver_data)
        await _clear_status_message(int(telegram_id))
        await driver_bot_instance.send_message(
            chat_id=int(telegram_id),
            text=DRIVER_L[lang]["approved"].format(fio=fio),
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"Не удалось уведомить курьера {telegram_id} об активации: {e}")


@dp.callback_query(F.data.startswith("reject_driver:"))
async def reject_driver_cb(callback: types.CallbackQuery):
    await callback.answer()
    if not await _is_manager(callback.from_user.id):
        return
    telegram_id = callback.data.split(":", 1)[1]
    fio = await asyncio.to_thread(_sync_reject_driver, telegram_id)
    if not fio:
        await callback.message.edit_text("❌ Курьер не найден или уже обработан.", reply_markup=None)
        return
    await callback.message.edit_text(
        f"❌ Курьер <b>{html.escape(fio)}</b> отклонён.",
        reply_markup=None, parse_mode="HTML"
    )
    try:
        driver_data = await asyncio.to_thread(_sync_get_driver, telegram_id)
        lang = _lang_from_driver_row(driver_data)
        await _clear_status_message(int(telegram_id))
        await driver_bot_instance.send_message(
            chat_id=int(telegram_id),
            text=DRIVER_L[lang]["rejected"],
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"Не удалось уведомить курьера {telegram_id} об отклонении: {e}")


# ─── Одобрение / отклонение смены ФИО курьера ───────────────────────────────

@dp.callback_query(F.data.startswith("napprove:"))
async def approve_name_change(callback: types.CallbackQuery):
    await callback.answer()
    if not await _is_manager(callback.from_user.id):
        return
    telegram_id = callback.data.split(":", 1)[1]
    result = await asyncio.to_thread(_sync_approve_name_change, telegram_id)
    if not result:
        await callback.message.edit_text("❌ Заявка не найдена или уже обработана.", reply_markup=None)
        return
    old_fio, new_fio = result
    await callback.message.edit_text(
        f"✅ ФИО курьера изменено: <b>{html.escape(old_fio)}</b> → <b>{html.escape(new_fio)}</b>",
        reply_markup=None, parse_mode="HTML"
    )
    try:
        await driver_bot_instance.send_message(
            chat_id=int(telegram_id),
            text=f"✅ <b>Ваше новое ФИО одобрено:</b> {html.escape(new_fio)}",
            parse_mode="HTML"
        )
    except Exception as e:
        logging.error(f"Не удалось уведомить курьера {telegram_id} об одобрении ФИО: {e}")


@dp.callback_query(F.data.startswith("nreject:"))
async def reject_name_change(callback: types.CallbackQuery):
    await callback.answer()
    if not await _is_manager(callback.from_user.id):
        return
    telegram_id = callback.data.split(":", 1)[1]
    rejected_fio = await asyncio.to_thread(_sync_reject_name_change, telegram_id)
    if not rejected_fio:
        await callback.message.edit_text("❌ Заявка не найдена или уже обработана.", reply_markup=None)
        return
    await callback.message.edit_text(
        f"❌ Заявка на ФИО «{html.escape(rejected_fio)}» отклонена.",
        reply_markup=None, parse_mode="HTML"
    )
    try:
        await driver_bot_instance.send_message(
            chat_id=int(telegram_id),
            text="❌ <b>Заявка на смену ФИО отклонена менеджером.</b>",
            parse_mode="HTML"
        )
    except Exception as e:
        logging.error(f"Не удалось уведомить курьера {telegram_id} об отклонении ФИО: {e}")


